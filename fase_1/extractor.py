#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extractor.py — Extractor acumulativo de listas de precios
Boutique Zepeda · books-label · Fase 1

Uso:
    python3 fase_1/extractor.py --config fase_1/config/config_ps.json
    python3 fase_1/extractor.py --config fase_1/config/config_pakar.json
    python3 fase_1/extractor.py --config fase_1/config/config_cklass.json

Proveedores soportados:
    PS     → Price Shoes   (encoding propietario OR texto limpio, auto-detectado)
    Pakar  → Pakar         (texto limpio, extracción tabular)
    Cklass → Cklass        (texto limpio, multi-catálogo en un solo PDF)
    Otro   → genérico      (texto limpio, offset configurable)

Dependencias:
    pdfplumber, pandas, openpyxl

─────────────────────────────────────────────────────────────────
Estructura base_precios.xlsx  (pestaña por proveedor, acumulativa)
─────────────────────────────────────────────────────────────────
PS / Pakar:
    catalogo | temp | pag | id | precio_base | redondea | precio_venta | fecha

Cklass:
    catalogo | temp | pag | modelo | precio_base | redondea | precio_venta | fecha

Nota: Cklass usa `modelo` en lugar de `id` porque el campo es texto libre
      ("DUO 417", "020-96", "SIX 503", "P28099", "Combo 040").
─────────────────────────────────────────────────────────────────
config_cklass.json — campos relevantes:
    proveedor          : "Cklass"
    temporada          : "PV26"
    pdf_input          : "lista_cruda/lista_cklass_pv26.pdf"
    catalogos_manuales : {
        "Dama":           [2, 3],        ← páginas PDF (1-based)
        "Gala & Glamour": [4],
        "Urban":          [4, 5, 6],
        ...
    }
    Los catálogos cuyos datos aparecen en páginas con marcador de texto
    ("C O L E C C I Ó N [NOMBRE]") son re-etiquetados automáticamente
    cuando el marcador se detecta dentro de la página.
─────────────────────────────────────────────────────────────────
"""

import math
import os
import re
import sys
import json
import logging
import argparse
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# ─────────────────────────────────────────────
#  Rutas fijas
# ─────────────────────────────────────────────

TABULADOR_PATH    = os.path.expanduser("~/books-label/fase_1/lista_cruda/tabulador.xlsx")
FASE2_PRECIOS_DIR = os.path.expanduser("~/books-label/fase_2/precios")
INTEGRACION_PATH  = os.path.expanduser("~/books-label/fase_1/salida/base_precios.xlsx")

# Columnas para PS / Pakar
COLUMNAS_STD    = ["catalogo", "temp", "pag", "id",     "precio_base", "redondea", "precio_venta", "fecha"]
# Columnas para Cklass (modelo reemplaza id, se agrega clave)
COLUMNAS_CKLASS = ["catalogo", "temp", "pag", "modelo", "precio_base", "redondea", "precio_venta", "fecha"]


# ─────────────────────────────────────────────
#  Helpers precio_venta
# ─────────────────────────────────────────────

def _round_excel(value: float, digits: int) -> float:
    """ROUND con redondeo 'half away from zero', igual que Excel."""
    if digits < 0:
        factor = 10 ** (-digits)
        return math.floor(float(value) / factor + 0.5) * factor
    factor = 10 ** digits
    return math.floor(float(value) * factor + 0.5) / factor


def _cargar_tabulador(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, header=0)
    df.columns = ["desde", "hasta", "sumar"]
    df = df.dropna(subset=["desde"]).copy()
    df["desde"] = pd.to_numeric(df["desde"], errors="coerce")
    df["sumar"] = pd.to_numeric(df["sumar"], errors="coerce")
    return df.sort_values("desde").reset_index(drop=True)


def _calcular_pv(precio_base, tab: pd.DataFrame):
    """Replica: =ROUND(IF(pb<200, ROUND(pb,-1)*1.5, pb+VLOOKUP(pb,tab,3,1)), -1)"""
    try:
        pb = float(precio_base)
    except (TypeError, ValueError):
        return None
    redondea = _round_excel(pb, -1)
    if pb < 200:
        return int(_round_excel(redondea * 1.5, -1))
    mask = tab["desde"] <= pb
    if not mask.any():
        return None
    sumar = float(tab.loc[mask.values.nonzero()[0][-1], "sumar"])
    return int(_round_excel(pb + sumar, -1))


# ─────────────────────────────────────────────
#  Argumentos
# ─────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Extractor acumulativo de listas de precios — Boutique Zepeda"
    )
    parser.add_argument("--config", required=True, help="Ruta al archivo JSON de configuración")
    return parser.parse_args()


# ─────────────────────────────────────────────
#  Logger
# ─────────────────────────────────────────────

def setup_logger(nombre: str, base_dir: str) -> logging.Logger:
    logger = logging.getLogger("extractor_01")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    return logger


# ─────────────────────────────────────────────
#  Decodificador Price Shoes
# ─────────────────────────────────────────────

def _decodificar(txt: str, offset: int) -> str:
    resultado = []
    i = 0
    while i < len(txt):
        m = re.match(r'\(cid:(\d+)\)', txt[i:])
        if m:
            n = int(m.group(1))
            if 19 <= n <= 28:
                resultado.append(str(n - 19))
            i += len(m.group(0))
        else:
            c = txt[i]
            o = ord(c)
            if offset != 0 and 32 <= o <= 126:
                n2 = o + offset
                resultado.append(chr(n2) if 32 <= n2 <= 126 else c)
            else:
                resultado.append(c)
            i += 1
    return ''.join(resultado)


def _tiene_encoding_ps(words: list) -> bool:
    return any('(cid:' in w['text'] for w in words)


# ─────────────────────────────────────────────
#  Limpieza de precio
# ─────────────────────────────────────────────

def _limpiar_precio(val: str) -> str | None:
    """
    Acepta todos los formatos de precio presentes en el PDF Cklass:

        "$579.00"        → "579"   (Dama, Urban, calzado — decimal con punto)
        "$1,079.00"      → "1079"  (miles con coma, decimal con punto)
        "349,00"         → "349"   (Ropa Caballero / Fashionline / SportBrands — decimal con coma)
        "$1.849,00"      → "1849"  (SportBrands miles con punto, decimal con coma)
        "$ 69"           → "69"    (Home / WE Cosmetics — entero con espacio tras $)
        "$1.199,00"      → "1199"
        "1.799,00"       → "1799"

    Descarta cualquier valor que contenga letras (ej. "Precio a tu conveniencia").
    """
    if not val:
        return None
    s = str(val).strip()
    # Descartar si tiene letras
    if re.search(r'[a-zA-ZáéíóúÁÉÍÓÚñÑ]', s):
        return None
    # Eliminar $ y espacios internos
    s = re.sub(r'[$\s]', '', s)
    if not s:
        return None

    # ── Normalizar a entero ──────────────────────────────────────────
    # Caso 1: Miles con punto Y decimal con coma: "1.849,00" o "1.799,00"
    if re.match(r'^\d{1,3}(\.\d{3})+,\d{2}$', s):
        s = s.replace('.', '').split(',')[0]
    # Caso 2: Miles con coma Y decimal con punto: "1,079.00"
    elif re.match(r'^\d{1,3}(,\d{3})+\.\d+$', s):
        s = s.replace(',', '').split('.')[0]
    # Caso 3: Decimal con coma sin miles: "349,00"
    elif re.match(r'^\d+,\d{2}$', s):
        s = s.split(',')[0]
    # Caso 4: Decimal con punto: "579.00" o "1199.00"
    elif re.match(r'^\d+\.\d+$', s):
        s = s.split('.')[0]
    # Caso 5: Entero puro: "69", "199", "1199"
    # (ya es correcto)

    if re.match(r'^\d+$', s) and s:
        return s
    return None


# ─────────────────────────────────────────────
#  Extractor Price Shoes
# ─────────────────────────────────────────────

class ExtractorPS:
    """
    Price Shoes: soporta AMBOS formatos del mismo proveedor:
      - PDFs con encoding propietario (cid: / offset ASCII)
      - PDFs con texto limpio y seleccionable
    La detección es automática por página.
    """

    _RE_ID = re.compile(r'^\d{6,8}$')

    def __init__(self, config: dict, logger: logging.Logger):
        self.config   = config
        self.logger   = logger
        self.offset   = config.get("encoding_offset", 29)
        self.tol_x    = config.get("tolerancia_x", 20.0)
        self.col_pag  = config.get("col_pag",    "Pag")
        self.col_id   = config.get("col_id",     "ID")
        self.col_prec = config.get("col_precio", "Sug_credito")
        self._merge_y = config.get("merge_y", 12)

    def extraer(self, pdf_path: str) -> pd.DataFrame:
        registros   = []
        col_x       = None
        ids_pdf_col = 0

        registros_por_pagina = []
        with pdfplumber.open(pdf_path) as pdf:
            self.logger.info(f"📄 Total páginas: {len(pdf.pages)}")
            self.logger.info("═" * 45)

            for i, page in enumerate(pdf.pages, 1):
                raw = page.extract_words(x_tolerance=3, y_tolerance=5, keep_blank_chars=False)
                if not raw:
                    continue

                if _tiene_encoding_ps(raw):
                    words = [{**w, "text": _decodificar(w["text"], self.offset)} for w in raw]
                    self.logger.debug(f"  Pág {i}: encoding PS → offset={self.offset}")
                else:
                    words = raw
                    self.logger.debug(f"  Pág {i}: texto limpio")

                nuevo_cx = self._detectar_encabezado(words)
                if nuevo_cx:
                    col_x = nuevo_cx
                    self.logger.debug(
                        f"  Pág {i}: encabezado OK — "
                        f"Pag@x={col_x['pag']:.1f}  ID@x={col_x['id']:.1f}  Precio@x={col_x['precio']:.1f}"
                    )

                if not col_x:
                    self.logger.warning(f"  Pág {i}: encabezado NO encontrado — omitida")
                    continue

                x_id = col_x["id"]
                for w in words:
                    tok = w["text"].strip()
                    x0_ = w["x0"]
                    x1_ = w.get("x1", x0_)
                    near_id = (abs(x0_ - x_id) <= self.tol_x or
                               (x0_ < x_id and x_id <= x1_ + self.tol_x))
                    if not near_id:
                        continue
                    if self._RE_ID.match(tok):
                        ids_pdf_col += 1
                    else:
                        m = re.search(r'(\d{6,8})$', tok)
                        if m:
                            ids_pdf_col += 1

                filas = self._extraer_filas(words, col_x)
                antes = len(registros)
                for fila in filas:
                    registros.append({
                        "pag":         fila.get("pag", "").strip(),
                        "id":          fila.get("id",  "").strip(),
                        "precio_base": _limpiar_precio(fila.get("precio", "")),
                    })
                df_pag = self._filtrar(pd.DataFrame(registros[antes:]))
                validos = len(df_pag)
                registros_por_pagina.append((i, validos))
                self.logger.info(f"  Pág {i}: {validos} registros")

        if registros_por_pagina:
            conteos   = [c for _, c in registros_por_pagina]
            promedio  = sum(conteos) / len(conteos)
            desviacion = (sum((c - promedio) ** 2 for c in conteos) / len(conteos)) ** 0.5
            x         = self.config.get("desviacion_alerta", 2)
            alertas   = [(pag, cnt) for pag, cnt in registros_por_pagina
                         if abs(cnt - promedio) > desviacion * x]
            self.logger.info("═" * 45)
            self.logger.info(f"📊 Promedio por página: {promedio:.1f}  σ={desviacion:.1f}")
            if alertas:
                detalle = ", ".join(f"Pág {pag} ({cnt} registros)" for pag, cnt in alertas)
                self.logger.warning(f"  ⚠️  Páginas fuera del rango: {detalle}")
            else:
                self.logger.info("  ✔ Todas las páginas dentro del rango esperado")

        df = pd.DataFrame(registros) if registros else pd.DataFrame(columns=["pag", "id", "precio_base"])
        df = self._filtrar(df)
        df.attrs["ids_esperados_pdf"] = ids_pdf_col
        return df

    def _detectar_encabezado(self, words: list) -> dict | None:
        filas_y = {}
        for w in words:
            filas_y.setdefault(round(w["top"]), []).append(w)

        ys_sorted = sorted(filas_y.keys())

        for y, tokens in sorted(filas_y.items()):
            textos = [t["text"].strip() for t in tokens]
            if self.col_pag not in textos or self.col_id not in textos:
                continue

            col_x = {}
            for t in tokens:
                txt = t["text"].strip()
                if txt == self.col_pag:
                    col_x["pag"] = t["x0"]
                elif txt == self.col_id:
                    col_x["id"] = t["x0"]
                elif txt == self.col_prec:
                    col_x["precio"] = t["x0"]

            if "precio" not in col_x:
                for y2 in ys_sorted:
                    if abs(y2 - y) == 0 or abs(y2 - y) > 5:
                        continue
                    for t in filas_y[y2]:
                        if t["text"].strip() == self.col_prec:
                            col_x["precio"] = t["x0"]
                            break

            if "pag" in col_x and "id" in col_x and "precio" in col_x:
                return col_x

        return None

    def _extraer_filas(self, words: list, col_x: dict) -> list:
        x_pag    = col_x["pag"]
        x_id     = col_x["id"]
        x_precio = col_x["precio"]
        tol      = self.tol_x

        filas_raw = {}
        for w in words:
            x     = w["x0"]
            x1_   = w.get("x1", x)
            token = w["text"].strip()
            if not token:
                continue
            if abs(x - x_pag) <= tol:
                col = "pag"
            elif (abs(x - x_id) <= tol or
                  (x < x_id and x_id <= x1_ + tol)):
                col = "id"
                if not self._RE_ID.match(token):
                    m = re.search(r'(\d{6,8})$', token)
                    if m:
                        token = m.group(1)
                    else:
                        continue
            elif abs(x - x_precio) <= tol:
                col = "precio"
            else:
                continue
            y_key = round(w["top"])
            filas_raw.setdefault(y_key, {})
            prev = filas_raw[y_key].get(col, "")
            filas_raw[y_key][col] = (prev + " " + token).strip() if prev else token

        bloques = []
        ys = sorted(filas_raw.keys())
        if not ys:
            return []

        bloque = dict(filas_raw[ys[0]])
        y_ini  = ys[0]

        for y in ys[1:]:
            if y - y_ini <= self._merge_y:
                for col, val in filas_raw[y].items():
                    if col not in bloque:
                        bloque[col] = val
                    else:
                        cur = bloque[col]
                        if col == "id" and self._RE_ID.match(val) and not self._RE_ID.match(cur):
                            bloque[col] = val
                        elif col == "precio" and re.match(r'^\$?\d', val) and not re.match(r'^\$?\d', cur):
                            bloque[col] = val
                        elif col == "pag" and re.match(r'^\d+$', val) and not re.match(r'^\d+$', cur):
                            bloque[col] = val
            else:
                bloques.append(bloque)
                bloque = dict(filas_raw[y])
                y_ini  = y

        bloques.append(bloque)
        return bloques

    def _filtrar(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        df = df[df["id"].astype(str).str.match(r"^\d{6,8}$")].copy()
        df = df[df["precio_base"].notna()].copy()
        df = df[~df["precio_base"].astype(str).str.contains(r'[a-zA-Z]', na=False)].copy()
        return df.reset_index(drop=True)


# ─────────────────────────────────────────────
#  Extractor Pakar
# ─────────────────────────────────────────────

class ExtractorPakar:
    """Pakar: texto limpio, extracción tabular por coordenadas X."""

    def __init__(self, config: dict, logger: logging.Logger):
        self.config   = config
        self.logger   = logger
        self.tol_x    = config.get("tolerancia_x", 20.0)
        self.tol_col  = config.get("tol_col", 0)
        self.col_pag  = config.get("col_pag",    "PÁG.")
        self.col_id   = config.get("col_id",     "CÓDIGO")
        self.col_prec = config.get("col_precio", "2 PAGOS")
        self.offset   = config.get("encoding_offset", 0)

    def extraer(self, pdf_path: str) -> pd.DataFrame:
        registros = []
        col_x     = None

        registros_por_pagina = []
        with pdfplumber.open(pdf_path) as pdf:
            self.logger.info(f"📄 Total páginas: {len(pdf.pages)}")
            self.logger.info("═" * 45)

            for i, page in enumerate(pdf.pages, 1):
                words = page.extract_words(x_tolerance=self.tol_x, y_tolerance=5, keep_blank_chars=False)
                if not words:
                    continue
                if self.offset:
                    words = [{**w, "text": _decodificar(w["text"], self.offset)} for w in words]

                header_y = self._detectar_encabezado(words)
                if header_y is not None:
                    col_x = self._mapear_columnas(words, header_y)
                if not col_x:
                    continue

                filas = self._agrupar_filas(words, header_y, col_x)
                antes = len(registros)
                for fila in filas:
                    registros.append({
                        "pag":         fila.get("pag", "").strip(),
                        "id":          fila.get("id",  "").strip(),
                        "precio_base": _limpiar_precio(fila.get("precio", "")),
                    })
                df_pag = self._filtrar(pd.DataFrame(registros[antes:]))
                validos = len(df_pag)
                registros_por_pagina.append((i, validos))
                self.logger.info(f"  Página {i}: {validos} registros")

        if registros_por_pagina:
            conteos    = [c for _, c in registros_por_pagina]
            promedio   = sum(conteos) / len(conteos)
            desviacion = (sum((c - promedio) ** 2 for c in conteos) / len(conteos)) ** 0.5
            x          = self.config.get("desviacion_alerta", 2)
            alertas    = [(pag, cnt) for pag, cnt in registros_por_pagina
                          if abs(cnt - promedio) > desviacion * x]
            self.logger.info("═" * 45)
            self.logger.info(f"📊 Promedio por página: {promedio:.1f}  σ={desviacion:.1f}")
            if alertas:
                detalle = ", ".join(f"Pág {pag} ({cnt} filas)" for pag, cnt in alertas)
                self.logger.warning(f"  ⚠️  Páginas fuera del rango: {detalle}")
            else:
                self.logger.info("  ✔ Todas las páginas dentro del rango esperado")

        df = pd.DataFrame(registros) if registros else pd.DataFrame(columns=["pag", "id", "precio_base"])
        return self._filtrar(df)

    def _detectar_encabezado(self, words):
        for w in words:
            txt = w["text"].strip()
            if txt == self.col_pag or self.col_pag in txt:
                return w["top"]
        return None

    def _mapear_columnas(self, words, header_y):
        col_x         = {}
        tol_y         = 15
        mapa          = {self.col_pag: "pag", self.col_id: "id", self.col_prec: "precio"}
        zona          = [w for w in words if abs(w["top"] - header_y) < tol_y]
        zona_por_fila = {}
        for w in zona:
            zona_por_fila.setdefault(round(w["top"]), []).append(w)
        for _, fila_tokens in zona_por_fila.items():
            for i, w in enumerate(fila_tokens):
                txt = w["text"].strip()
                for nombre, alias in mapa.items():
                    if alias in col_x:
                        continue
                    if txt == nombre or nombre in txt:
                        col_x[alias] = w["x0"]
                        break
                    if i + 1 < len(fila_tokens):
                        txt2 = txt + " " + fila_tokens[i + 1]["text"].strip()
                        if txt2 == nombre or nombre in txt2:
                            col_x[alias] = w["x0"]
                            break
        return col_x

    def _agrupar_filas(self, words, header_y, col_x):
        filas   = {}
        aliases = list(col_x.keys())
        xs      = [col_x[a] for a in aliases]
        for w in words:
            if w["top"] <= header_y + 5:
                continue
            y_key = round(w["top"])
            token = w["text"].strip()
            if not token:
                continue
            idx = min(range(len(xs)), key=lambda k: abs(w["x0"] - xs[k]))
            if self.tol_col > 0 and abs(w["x0"] - xs[idx]) > self.tol_col:
                continue
            col = aliases[idx]
            filas.setdefault(y_key, {})
            prev = filas[y_key].get(col, "")
            filas[y_key][col] = (prev + " " + token).strip() if prev else token
        return list(filas.values())

    def _filtrar(self, df):
        if df.empty:
            return df
        df = df[df["precio_base"].notna()].copy()
        df = df[df["id"].str.strip() != ""].copy()
        df = df[~df["precio_base"].astype(str).str.contains(r'[a-zA-Z]', na=False)].copy()
        return df.reset_index(drop=True)


# ─────────────────────────────────────────────
#  Extractor Cklass
# ─────────────────────────────────────────────

# Mapa de marcadores de texto detectables en el PDF → nombre normalizado del catálogo.
# Patrón en el PDF: "C O N F O R T" (anterior a la línea COLECCIÓN) o
#                   "C O L E C C I Ó N C A L Z A D O N I Ñ A S ..."
# Se colapsan los espacios intermedios para comparar.
_CKLASS_MARCADORES = {
    # texto colapsado (sin espacios internos) → nombre canónico
    "CONFORT":              "Confort",
    "SIXYDÚOPACK":          "Six & Duo Pack",
    "SIXYDUOPACK":          "Six & Duo Pack",
    "CALZADONIÑAS":         "Calzado Niñas",
    "ROPANIÑAS":            "Ropa Niñas",
    "CALZADONIÑOS":         "Calzado Niños",
    "ROPANIÑOS":            "Ropa Niños",
    "ROPANICABALLERO":      "Caballeros",
    "CABALLERO":            "Caballeros",
    "ROPACABALLERO":        "Caballeros",
    "ROPAFASHIONLINE":      "Ropa Fashion Online",
    "FASHIONLINE":          "Ropa Fashion Online",
    "JOYERÍA":              "Joyería",
    "JOYERIA":              "Joyería",
    "LENCERÍA":             "Lenceria",
    "LENCERIA":             "Lenceria",
    "BOLSOSHANDBAGS":       "Bolsos Handbags",
    "HANDBAGS":             "Bolsos Handbags",
    "WECOSMETICS":          "WE Cosmetics",
    "COSMETICS":            "WE Cosmetics",
    "HOME":                 "Home",
    "SPORTBRANDSDAMA":      "SportBrands Dama",
    "SPORTBRANDSCABALLERO": "SportBrands Caballero",
    "ACCESORIOS":           "Accesorios",
    "SUPLEMENTOS":          "Suplementos",
    "SPORTBRANDSKIDS":      "SportBrands Kids & Teens",
    "SPORTBRANDSTEENS":     "SportBrands Kids & Teens",
    "KIDSTEENS":            "SportBrands Kids & Teens",
}


def _colapsar(txt: str) -> str:
    """Elimina todos los espacios de una cadena y la pone en mayúsculas."""
    return re.sub(r'\s+', '', txt).upper()


def _detectar_marcador_catalogo(lines: list[str]) -> str | None:
    """
    Busca en las líneas de texto de una página si aparece un marcador
    reconocible de catálogo ANTES de la línea COLECCIÓN o suelto.
    Devuelve el nombre canónico o None.
    """
    for line in lines:
        collapsed = _colapsar(line)
        # Quitar prefijo "COLECCIÓN" si está pegado
        collapsed_clean = re.sub(r'^C[OÓ]LECCI[OÓ]N', '', collapsed)
        # Quitar sufijos de temporada
        collapsed_clean = re.sub(r'PRIMAVERAV[EÉ]RANO\d{0,4}$', '', collapsed_clean)
        collapsed_clean = collapsed_clean.strip()
        for key, nombre in _CKLASS_MARCADORES.items():
            if collapsed_clean == key or collapsed_clean.startswith(key):
                return nombre
    return None


class ExtractorCklass:
    """
    Cklass: un único PDF contiene múltiples catálogos.

    Estrategia de asignación de catálogo (por página PDF):
    1. Se construye un dict  pag_pdf → nombre_catalogo  a partir de
       `catalogos_manuales` del JSON (cada catálogo declara en qué
       páginas PDF aparecen sus registros).
    2. Cuando una misma página PDF aparece en varios catálogos
       (transición en medio de página), el extractor detecta el marcador
       textual "C O L E C C I Ó N [NOMBRE]" dentro de la página para
       saber en qué línea cambia el catálogo activo.
    3. Filas sin precio numérico ("Precio a tu conveniencia") se descartan.

    Columnas de salida: pag | modelo | precio_base
    (la clave de 7 dígitos NO se exporta — es un campo interno del PDF)
    """

    # Encabezados esperados en el PDF
    COL_PAG    = "PÁGINA"
    COL_MODELO = "MODELO"
    COL_CLAVE  = "CLAVE"
    COL_PRECIO = "CRÉDITO"

    # Tolerancia horizontal para alinear tokens a columnas (px)
    TOL_X = 25.0

    def __init__(self, config: dict, logger: logging.Logger):
        self.config  = config
        self.logger  = logger
        # catalogos_manuales: {nombre_catalogo: [pag_pdf, ...]}
        self.cat_map: dict[str, list[int]] = config.get("catalogos_manuales", {})
        # Invertir: pag_pdf → [nombre_catalogo, ...]  (puede haber varios)
        self._pag_a_cats: dict[int, list[str]] = {}
        for nombre, paginas in self.cat_map.items():
            for p in paginas:
                self._pag_a_cats.setdefault(p, []).append(nombre)

    # ── Interfaz pública ──────────────────────────────────────────────

    def extraer(self, pdf_path: str) -> pd.DataFrame:
        """
        Retorna DataFrame con columnas: catalogo | pag | modelo | precio_base
        """
        todos: list[dict] = []

        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            self.logger.info(f"📄 Total páginas: {total}")
            self.logger.info("═" * 45)

            # col_x persiste entre páginas (el encabezado puede estar
            # en una página imagen sin texto pero las posiciones X son
            # consistentes en todo el PDF)
            col_x: dict | None = None

            for pdf_pag_idx, page in enumerate(pdf.pages, 1):
                cats_en_pag = self._pag_a_cats.get(pdf_pag_idx, [])
                if not cats_en_pag:
                    # Página no declarada en el config → saltar silenciosamente
                    continue

                words = page.extract_words(x_tolerance=3, y_tolerance=3,
                                           keep_blank_chars=False)
                if not words:
                    self.logger.debug(f"  Pág {pdf_pag_idx}: imagen/sin texto — omitida")
                    continue

                # Detectar/refrescar posiciones X del encabezado
                # Puede haber múltiples encabezados en la misma página
                # (ej. pág 23: Fashionline + Joyería + Lencería)
                encabezados = self._detectar_todos_encabezados(words)
                if encabezados:
                    col_x = encabezados[0][1]  # primer encabezado de la página
                    if len(encabezados) > 1:
                        self.logger.debug(
                            f"  Pág {pdf_pag_idx}: {len(encabezados)} encabezados detectados"
                        )
                    else:
                        self.logger.debug(
                            f"  Pág {pdf_pag_idx}: encabezado detectado — "
                            f"pag@x={col_x['pag']:.0f}  modelo@x={col_x['modelo']:.0f}  "
                            f"precio@x={col_x['precio']:.0f}"
                        )

                if col_x is None:
                    self.logger.warning(f"  Pág {pdf_pag_idx}: sin encabezado conocido — omitida")
                    continue

                # Extraer filas brutas con soporte para múltiples encabezados
                filas_brutas = self._extraer_filas_brutas(words, encabezados if encabezados else [(0.0, col_x)])

                # Si la página pertenece a UN solo catálogo, asignación directa
                if len(cats_en_pag) == 1:
                    cat_actual = cats_en_pag[0]
                    registros = self._asignar_catalogo_unico(
                        filas_brutas, cat_actual, pdf_pag_idx
                    )
                else:
                    # Página de transición: detectar cambio de catálogo por marcador
                    registros = self._asignar_catalogo_transicion(
                        words, filas_brutas, cats_en_pag, pdf_pag_idx
                    )

                todos.extend(registros)
                validos = sum(1 for r in registros if r["precio_base"] is not None)
                self.logger.info(
                    f"  Pág {pdf_pag_idx:3d}: {validos:3d} registros "
                    f"[{', '.join(cats_en_pag)}]"
                )

        df = pd.DataFrame(todos) if todos else pd.DataFrame(
            columns=["catalogo", "pag", "modelo", "precio_base"]
        )
        df = self._filtrar(df)

        # Estadísticas por catálogo
        self.logger.info("═" * 45)
        for cat, grp in df.groupby("catalogo", sort=False):
            self.logger.info(f"  {cat:30s}: {len(grp):4d} registros")
        self.logger.info(f"  {'TOTAL':30s}: {len(df):4d} registros")

        return df

    # ── Detección de encabezado ───────────────────────────────────────

    def _detectar_columnas(self, words: list) -> dict | None:
        """
        Detecta el PRIMER encabezado con PÁGINA + MODELO + CRÉDITO y
        devuelve sus posiciones X.  Se llama para el primer encabezado;
        _detectar_todos_encabezados maneja páginas con varios.
        """
        return self._detectar_todos_encabezados(words)[0][1] if self._detectar_todos_encabezados(words) else None

    def _detectar_todos_encabezados(self, words: list) -> list[tuple[float, dict]]:
        """
        Detecta todos los encabezados de tabla en la página.
        Solo busca las tres columnas que siempre están presentes:
            PÁGINA → pag
            MODELO → modelo
            CRÉDITO → precio

        Devuelve lista de (y_top, col_x) ordenada ascendente.
        Las posiciones X se calibran desde el encabezado real de cada sub-tabla,
        lo que cubre tanto el formato estándar (con CLAVE) como los formatos
        alternativos (Joyería, Bolsos Handbags, WE Cosmetics, Home, SportBrands).
        """
        filas_y: dict[int, list] = {}
        for w in words:
            if w["x0"] < 0:
                continue
            filas_y.setdefault(round(w["top"]), []).append(w)

        resultado = []
        for y in sorted(filas_y):
            row_texts = {w["text"].strip(): w for w in filas_y[y]}
            if self.COL_PAG not in row_texts or self.COL_MODELO not in row_texts:
                continue

            col_x: dict = {}
            for w in filas_y[y]:
                txt = w["text"].strip()
                if txt == self.COL_PAG:
                    col_x["pag"]    = w["x0"]
                elif txt == self.COL_MODELO:
                    col_x["modelo"] = w["x0"]
                elif txt == self.COL_PRECIO:
                    col_x["precio"] = w["x0"]

            # CRÉDITO puede estar en fila adyacente (encabezado partido en 2 líneas)
            if "precio" not in col_x:
                for y2 in sorted(filas_y):
                    if y2 == y or abs(y2 - y) > 8:
                        continue
                    for w in filas_y[y2]:
                        if w["text"].strip() == self.COL_PRECIO:
                            col_x["precio"] = w["x0"]
                            break
                    if "precio" in col_x:
                        break

            if "pag" in col_x and "modelo" in col_x and "precio" in col_x:
                resultado.append((float(y), col_x))

        return resultado

    # ── Extracción de filas brutas ────────────────────────────────────

    def _extraer_filas_brutas(self, words: list, encabezados: list[tuple[float, dict]]) -> list[dict]:
        """
        Agrupa tokens por fila (y) y los asigna a la columna más cercana
        (pag | modelo | precio).

        Estrategia de asignación de precio (en orden de precedencia):
        1. Token cuyo x0 está dentro de TOL_X del x_precio del encabezado vigente.
        2. Si no hay candidato por posición, el token numérico más a la derecha
           de la fila (los precios siempre son los valores más derechos antes de
           TALLA/NUMERACIÓN). Esto cubre tablas sin CLAVE donde el encabezado
           "CRÉDITO" tiene un offset respecto a los datos reales.

        Tokens con x < 0 se descartan (sub-tablas doble columna en SportBrands).
        """
        enc_sorted = sorted(encabezados, key=lambda t: t[0])

        def _get_col_x(y_fila: float) -> dict:
            active = enc_sorted[0][1]
            for y_enc, cx in enc_sorted:
                if y_fila >= y_enc:
                    active = cx
                else:
                    break
            return active

        tol    = self.TOL_X
        # RE para reconocer un token de precio numérico
        _RE_PRECIO = re.compile(
            r'^\$?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?$'
        )

        # Acumular por y_round → {pag, modelo, _precio_candidatos: [(x, tok)]}
        acumulados: dict[int, dict] = {}

        for w in words:
            tok = w["text"].strip()
            if not tok:
                continue
            x = w["x0"]
            y = round(w["top"])

            if x < 0:
                continue

            # Ignorar encabezados y decorativos
            if tok in (self.COL_PAG, self.COL_MODELO, self.COL_CLAVE,
                       self.COL_PRECIO, "COLOR", "CONTADO", "NUMERACIÓN",
                       "TALLA", "OBSERVACIÓN"):
                continue
            # Ignorar '$' suelto: WE Cosmetics/Home separan "$ 69" en dos tokens
            if tok == "$":
                continue
            if re.search(r'C\s*O\s*L\s*E\s*C\s*C', tok, re.IGNORECASE):
                continue
            if re.search(r'PRIMAVERA|VERANO', tok, re.IGNORECASE):
                continue

            col_x    = _get_col_x(float(w["top"]))
            x_pag    = col_x["pag"]
            x_modelo = col_x["modelo"]
            x_precio = col_x["precio"]

            dist_pag    = abs(x - x_pag)
            dist_modelo = abs(x - x_modelo)
            dist_precio = abs(x - x_precio)

            acumulados.setdefault(y, {"y_top": w["top"], "_precio_cands": []})

            if dist_pag <= tol and dist_pag == min(dist_pag, dist_modelo, dist_precio):
                prev = acumulados[y].get("pag", "")
                acumulados[y]["pag"] = (prev + " " + tok).strip() if prev else tok

            elif dist_precio <= tol:
                # En rango exacto del encabezado → precio seguro
                acumulados[y]["_precio_cands"].append((dist_precio, x, tok))

            elif dist_modelo <= tol * 2.5:
                prev = acumulados[y].get("modelo", "")
                acumulados[y]["modelo"] = (prev + " " + tok).strip() if prev else tok

            elif _RE_PRECIO.match(tok):
                # Token numérico fuera de rango de posición →
                # candidato de precio por valor (para tablas sin CLAVE)
                acumulados[y]["_precio_cands"].append((dist_precio, x, tok))

        # Resolver precio: tomar el candidato con menor distancia a x_precio
        filas = []
        for y, data in acumulados.items():
            cands = data.pop("_precio_cands", [])
            if cands:
                # Ordenar por distancia; si hay empate, elegir el más a la derecha (CRÉDITO)
                cands.sort(key=lambda c: (c[0], -c[1]))
                data["precio"] = cands[0][2]
            data.pop("_precio_dist", None)
            filas.append(data)

        filas.sort(key=lambda r: r["y_top"])
        return filas

    # ── Asignación de catálogo ────────────────────────────────────────

    def _asignar_catalogo_unico(
        self, filas: list[dict], catalogo: str, pdf_pag: int
    ) -> list[dict]:
        """Todas las filas de la página pertenecen al mismo catálogo."""
        resultado = []
        for fila in filas:
            precio = _limpiar_precio(fila.get("precio", ""))
            modelo = fila.get("modelo", "").strip()
            if not modelo:
                continue
            resultado.append({
                "catalogo":    catalogo,
                "pag":         fila.get("pag", "").strip(),
                "modelo":      modelo,
                "precio_base": precio,
                "_pdf_pag":    pdf_pag,
            })
        return resultado

    def _asignar_catalogo_transicion(
        self,
        words: list,
        filas: list[dict],
        cats_en_pag: list[str],
        pdf_pag: int,
    ) -> list[dict]:
        """
        Página que contiene registros de DOS o más catálogos consecutivos.

        Estrategia:
        1. Buscar marcadores textuales "C O L E C C I Ó N [NOMBRE]"
           dentro de las líneas de la página y registrar su y_top.
        2. Iniciar con el primer catálogo de la lista declarada en el config.
        3. Cuando una fila supera el y_top de un marcador conocido,
           cambiar al catálogo correspondiente.
        """
        # Paso 1: extraer texto de la página en líneas
        texto_page = " ".join(w["text"] for w in words)
        lines_page = []
        y_lines: dict[int, str] = {}
        for w in words:
            y = round(w["top"])
            y_lines[y] = y_lines.get(y, "") + " " + w["text"]

        # Paso 2: detectar posiciones y de marcadores
        # Estructura: [(y_top_marcador, nombre_catalogo), ...]
        cambios: list[tuple[float, str]] = []
        for y_key in sorted(y_lines):
            line = y_lines[y_key].strip()
            collapsed = _colapsar(line)
            for key, nombre in _CKLASS_MARCADORES.items():
                if collapsed == key or key in collapsed:
                    if nombre in cats_en_pag:
                        cambios.append((float(y_key), nombre))
                        break

        # Eliminar duplicados y ordenar
        vistos = set()
        cambios_uniq = []
        for item in sorted(cambios, key=lambda t: t[0]):
            if item[1] not in vistos:
                cambios_uniq.append(item)
                vistos.add(item[1])
        cambios = cambios_uniq

        # Paso 3: catálogo inicial = primero de la lista del config
        cat_actual = cats_en_pag[0]
        cambio_idx = 0

        resultado = []
        for fila in filas:
            y_fila = fila["y_top"]

            # ¿Pasamos el umbral del siguiente catálogo?
            while cambio_idx < len(cambios) and y_fila >= cambios[cambio_idx][0]:
                cat_actual = cambios[cambio_idx][1]
                self.logger.debug(
                    f"    Transición en y={cambios[cambio_idx][0]:.0f} → '{cat_actual}'"
                )
                cambio_idx += 1

            precio = _limpiar_precio(fila.get("precio", ""))
            modelo = fila.get("modelo", "").strip()
            if not modelo:
                continue
            resultado.append({
                "catalogo":    cat_actual,
                "pag":         fila.get("pag", "").strip(),
                "modelo":      modelo,
                "precio_base": precio,
                "_pdf_pag":    pdf_pag,
            })

        return resultado

    # ── Filtrado final ────────────────────────────────────────────────

    def _filtrar(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        # Quitar filas sin precio válido
        df = df[df["precio_base"].notna()].copy()
        # Precio mínimo realista: descartar < 19 (ruido de claves/numeración)
        df["_pb_num"] = pd.to_numeric(df["precio_base"], errors="coerce")
        df = df[df["_pb_num"] >= 19].copy()
        df = df.drop(columns=["_pb_num"])
        # Quitar modelos vacíos, numéricos puros cortos (footers) o de 1-3 chars
        df = df[df["modelo"].astype(str).str.strip() != ""].copy()
        df = df[~df["modelo"].astype(str).str.match(r'^[\d\s]{1,4}$')].copy()
        # Quitar columna interna
        df = df.drop(columns=["_pdf_pag"], errors="ignore")
        return df.reset_index(drop=True)


# ─────────────────────────────────────────────
#  Escritura acumulativa en xlsx
# ─────────────────────────────────────────────

def _abrir_o_crear_xlsx(path: str):
    if os.path.isfile(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _obtener_o_crear_pestaña(wb, nombre: str, columnas: list):
    """Devuelve la hoja del proveedor; la crea con encabezado si no existe."""
    if nombre in wb.sheetnames:
        return wb[nombre]
    ws = wb.create_sheet(title=nombre)
    ws.append(columnas)
    return ws


def _estilizar_encabezado(ws):
    for cell in ws[1]:
        cell.fill      = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.font      = Font(name="Calibri", color="FFFFFF", bold=False)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")


def escribir_acumulativo_std(
    df: pd.DataFrame,
    config: dict,
    tab: pd.DataFrame,
    logger: logging.Logger,
):
    """Escritura para PS / Pakar (columna `id`)."""
    os.makedirs(os.path.dirname(INTEGRACION_PATH), exist_ok=True)
    proveedor = config.get("proveedor", "Otro").strip()
    catalogo  = config.get("catalogo", "")
    temporada = config.get("temporada", "")
    fecha     = datetime.now().strftime("%Y-%m-%d")

    wb = _abrir_o_crear_xlsx(INTEGRACION_PATH)
    ws = _obtener_o_crear_pestaña(wb, proveedor, COLUMNAS_STD)

    nuevos = 0
    for record in df.itertuples(index=False):
        try:
            pb = int(float(record.precio_base))
        except (TypeError, ValueError):
            pb = record.precio_base

        pv       = _calcular_pv(pb, tab) if isinstance(pb, (int, float)) else None
        redondea = int(_round_excel(pb, -1)) if isinstance(pb, (int, float)) else None

        ws.append([catalogo, temporada, record.pag, record.id,
                   pb, redondea, pv, fecha])
        nuevos += 1

    _estilizar_encabezado(ws)
    wb.save(INTEGRACION_PATH)
    logger.info(f"✅ {nuevos} registros → pestaña '{proveedor}' en {INTEGRACION_PATH}")
    return nuevos


def escribir_acumulativo_cklass(
    df: pd.DataFrame,
    config: dict,
    tab: pd.DataFrame,
    logger: logging.Logger,
):
    """
    Escritura para Cklass.
    - Pestaña fija: "Cklass"
    - El campo `catalogo` viene del DataFrame (por registro), NO del config.
    - Columnas: catalogo | temp | pag | modelo | precio_base | redondea | precio_venta | fecha
    - Acumulativo: los registros se agregan debajo de los existentes.
    - Cada llamada puede añadir registros de TODOS los catálogos del PDF.
    """
    os.makedirs(os.path.dirname(INTEGRACION_PATH), exist_ok=True)
    temporada = config.get("temporada", "")
    fecha     = datetime.now().strftime("%Y-%m-%d")

    wb = _abrir_o_crear_xlsx(INTEGRACION_PATH)
    ws = _obtener_o_crear_pestaña(wb, "Cklass", COLUMNAS_CKLASS)

    nuevos = 0
    for record in df.itertuples(index=False):
        try:
            pb = int(float(record.precio_base))
        except (TypeError, ValueError):
            pb = record.precio_base

        pv       = _calcular_pv(pb, tab) if isinstance(pb, (int, float)) else None
        redondea = int(_round_excel(pb, -1)) if isinstance(pb, (int, float)) else None

        ws.append([
            record.catalogo,  # A  catalogo  (por registro, no del config)
            temporada,        # B  temp
            record.pag,       # C  pag
            record.modelo,    # D  modelo    (texto libre)
            pb,               # E  precio_base
            redondea,         # F  redondea
            pv,               # G  precio_venta
            fecha,            # H  fecha
        ])
        nuevos += 1

    _estilizar_encabezado(ws)
    wb.save(INTEGRACION_PATH)
    logger.info(f"✅ {nuevos} registros → pestaña 'Cklass' en {INTEGRACION_PATH}")
    return nuevos


# ─────────────────────────────────────────────
#  Clase principal
# ─────────────────────────────────────────────

class ExtractorCatalogo:

    def __init__(self, config: dict, base_dir: str):
        nombre = os.path.splitext(
            os.path.basename(config.get("_config_path", "extractor"))
        )[0]
        self.logger    = setup_logger(nombre, base_dir)
        self.config    = config
        self.base_dir  = base_dir
        self.proveedor = config.get("proveedor", "PS").strip()

        self.pdf_path = os.path.join(base_dir, config.get("pdf_input", ""))
        os.makedirs(os.path.dirname(INTEGRACION_PATH), exist_ok=True)

        self.logger.info("═" * 45)
        self.logger.info(f"🏭 Proveedor:       {self.proveedor}")
        self.logger.info(f"📄 PDF de entrada:  {self.pdf_path}")
        self.logger.info(f"🛢️  Base Precios:    {INTEGRACION_PATH}")
        self.logger.info(f"📅 Temporada:       {config.get('temporada', '')}")
        if self.proveedor.upper() == "CKLASS":
            cats = list(config.get("catalogos_manuales", {}).keys())
            self.logger.info(f"📚 Catálogos ({len(cats)}): {', '.join(cats)}")

    def _factory(self):
        p = self.proveedor.upper()
        if p == "PS":
            self.logger.info("🔧 Modo: Price Shoes (auto-detect encoding + coordenadas X)")
            return ExtractorPS(self.config, self.logger)
        elif p == "PAKAR":
            self.logger.info("🔧 Modo: Pakar (texto limpio, tabular)")
            return ExtractorPakar(self.config, self.logger)
        elif p == "CKLASS":
            self.logger.info("🔧 Modo: Cklass (multi-catálogo, coordenadas X)")
            return ExtractorCklass(self.config, self.logger)
        elif p == "OTRO":
            self.logger.info("🔧 Modo: Pakar genérico (texto limpio, tabular)")
            return ExtractorPakar(self.config, self.logger)
        else:
            self.logger.warning(
                f"⚠️  Proveedor '{self.proveedor}' no reconocido — "
                "usa PS, Pakar, Cklass u Otro"
            )
            raise ValueError(f"Proveedor desconocido: {self.proveedor}")

    def ejecutar(self):
        if not os.path.isfile(self.pdf_path):
            self.logger.error(f"❌ PDF no encontrado: {self.pdf_path}")
            raise FileNotFoundError(self.pdf_path)

        if not os.path.isfile(TABULADOR_PATH):
            self.logger.error(f"❌ Tabulador no encontrado: {TABULADOR_PATH}")
            raise FileNotFoundError(TABULADOR_PATH)

        tab       = _cargar_tabulador(TABULADOR_PATH)
        extractor = self._factory()
        self.logger.info("🚀 Iniciando extracción...")
        df = extractor.extraer(self.pdf_path)

        if df.empty:
            self.logger.warning("⚠️  No se extrajeron registros.")
            self.logger.error("🔴 EXTRACCIÓN FALLIDA")
            return

        # ── Estadísticas generales ────────────────────────────────────
        n    = len(df)
        pmin = df["precio_base"].min()
        pmax = df["precio_base"].max()

        self.logger.info("═" * 45)
        self.logger.info(f"   Registros extraídos  : {n}")

        if self.proveedor.upper() == "CKLASS":
            u = df["modelo"].nunique()
            self.logger.info(f"   Modelos únicos       : {u}")
        else:
            u            = df["id"].nunique()
            esperados    = df.attrs.get("ids_esperados_pdf")
            self.logger.info(f"   IDs únicos           : {u}  {'(hay IDs repetidos)' if u < n else '(sin duplicados)'}")
            if esperados is not None:
                diff = esperados - n
                pct  = f"{n / esperados * 100:.1f}%"
                self.logger.info(f"   IDs en PDF (columna) : {esperados}")
                self.logger.info(f"   Cobertura            : {pct}  ({n}/{esperados})")
                if diff == 0:
                    self.logger.info("   ✔ COMPLETO — todos los registros capturados")
                elif diff > 0:
                    self.logger.warning(f"   ⚠ Faltan {diff} registro(s) — revisar manualmente")
                else:
                    self.logger.warning(f"   ⚠ Extraídos {-diff} de más — posibles duplicados")

        self.logger.info(f"   Rango precios        : ${pmin} – ${pmax}")
        self.logger.info("═" * 45)

        # ── Escritura en base_precios.xlsx ────────────────────────────
        if self.proveedor.upper() == "CKLASS":
            escribir_acumulativo_cklass(df, self.config, tab, self.logger)
        else:
            escribir_acumulativo_std(df, self.config, tab, self.logger)

        # ── fase_2/precios/ ───────────────────────────────────────────
        self._escribir_fase2(df, tab)

        self.logger.info(f"[STAT] proveedor={self.proveedor}")
        self.logger.info(f"[STAT] registros={n}")

    def _escribir_fase2(self, df: pd.DataFrame, tab: pd.DataFrame):
        nombre_base = os.path.splitext(
            os.path.basename(self.config.get("excel_output", "salida.xlsx"))
        )[0]
        fase2_path = os.path.join(FASE2_PRECIOS_DIR, f"{nombre_base}.xlsx")
        os.makedirs(FASE2_PRECIOS_DIR, exist_ok=True)

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Precios"

        # Para Cklass el identificador es modelo; para los demás es id
        if self.proveedor.upper() == "CKLASS":
            ws2.append(["catalogo", "modelo", "precio_venta"])
            for record in df.itertuples(index=False):
                try:
                    pb = int(float(record.precio_base))
                except (TypeError, ValueError):
                    pb = record.precio_base
                pv = _calcular_pv(pb, tab) if isinstance(pb, (int, float)) else None
                ws2.append([record.catalogo, record.modelo, pv])
        else:
            ws2.append(["ID", "precio_venta"])
            for record in df.itertuples(index=False):
                try:
                    pb = int(float(record.precio_base))
                except (TypeError, ValueError):
                    pb = record.precio_base
                pv = _calcular_pv(pb, tab) if isinstance(pb, (int, float)) else None
                ws2.append([record.id, pv])

        wb2.save(fase2_path)
        self.logger.info(f"✅ Excel (fase 2) generado: {fase2_path}")


# ─────────────────────────────────────────────
#  Punto de entrada
# ─────────────────────────────────────────────

if __name__ == "__main__":
    args        = parse_args()
    config_path = os.path.abspath(args.config)
    BASE        = os.path.dirname(os.path.abspath(__file__))

    if not os.path.isfile(config_path):
        print(f"❌ Archivo de configuración no encontrado: {config_path}")
        sys.exit(1)

    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)

    config["_config_path"] = config_path

    try:
        app = ExtractorCatalogo(config, BASE)
        app.ejecutar()
    except Exception as e:
        logging.getLogger("extractor_01").error(f"🔥 Error crítico: {e}", exc_info=True)
        sys.exit(1)
