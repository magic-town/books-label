#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extractor_01.py — Extractor acumulativo de listas de precios
Boutique Zepeda · books-label · Fase 1

Uso:
    python3 fase_1/extractor_01.py --config fase_1/config/config_ps.json

Cambios respecto a extractor.py:
    - Acumula registros en un xlsx único con pestañas por proveedor
    - Campos nuevos: name, temporada, fecha (automática)
    - No sobreescribe — encuentra la última fila y agrega
    - fase_2/precios/ no se toca (sin cambios)

Proveedores soportados:
    PS     → Price Shoes  (encoding propietario OR texto limpio, auto-detectado)
    Pakar  → Pakar        (texto limpio, extracción tabular)
    Cklass → Cklass       (texto limpio, precios con $ y decimales)
    Otro   → genérico     (texto limpio, offset configurable)

Dependencias Python (requirements.txt):
    pdfplumber, pandas, openpyxl

Estructura del xlsx acumulativo — fase_1/salidas/lista_precios.xlsx:
    Pestaña por proveedor (PS, Pakar, Cklass, Otro)
    Columnas:
        A  name          — del JSON
        B  temporada     — del JSON
        C  pag           — del PDF
        D  id            — del PDF (texto)
        E  precio_base   — del PDF
        F  redondea      — =ROUND(E,-1)
        G  precio_venta  — valor estático calculado con tabulador
        H  fecha         — datetime de ejecución (automático)

Segundo output — fase_2/precios/ (sin cambios):
    A  ID            valor estático
    B  precio_venta  valor estático
"""

import math
import os
import re
import sys
import json
import logging
import argparse
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook

# ─────────────────────────────────────────────
#  Rutas fijas
# ─────────────────────────────────────────────

TABULADOR_PATH    = os.path.expanduser("~/books-label/fase_1/lista_cruda/tabulador.xlsx")
FASE2_PRECIOS_DIR = os.path.expanduser("~/books-label/fase_2/precios")
INTEGRACION_PATH  = os.path.expanduser("~/books-label/fase_1/salidas/lista_precios.xlsx")

COLUMNAS = ["name", "temporada", "pag", "id", "precio_base",
            "redondea", "precio_venta", "fecha"]


# ─────────────────────────────────────────────
#  Helpers precio_venta
# ─────────────────────────────────────────────

def _round_excel(value: float, digits: int) -> float:
    """ROUND con redondeo 'half away from zero', igual que Excel."""
    if digits < 0:
        factor = 10 ** (-digits)
        return math.floor(float(value) / factor + 0.5) * factor
    factor = 10 ** digits
    return math.floor(float(value) * factor + 0.5) / factor


def _cargar_tabulador(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, header=0)
    df.columns = ["desde", "hasta", "sumar"]
    df = df.dropna(subset=["desde"]).copy()
    df["desde"] = pd.to_numeric(df["desde"], errors="coerce")
    df["sumar"] = pd.to_numeric(df["sumar"], errors="coerce")
    return df.sort_values("desde").reset_index(drop=True)


def _calcular_pv(precio_base, tab: pd.DataFrame):
    """Replica: =ROUND(IF(pb<200, ROUND(pb,-1)*1.5, pb+VLOOKUP(pb,tab,3,1)), -1)"""
    try:
        pb = float(precio_base)
    except (TypeError, ValueError):
        return None
    redondea = _round_excel(pb, -1)
    if pb < 200:
        return int(_round_excel(redondea * 1.5, -1))
    mask = tab["desde"] <= pb
    if not mask.any():
        return None
    sumar = float(tab.loc[mask.values.nonzero()[0][-1], "sumar"])
    return int(_round_excel(pb + sumar, -1))


# ─────────────────────────────────────────────
#  Argumentos
# ─────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Extractor acumulativo de listas de precios — Boutique Zepeda"
    )
    parser.add_argument("--config", required=True, help="Ruta al archivo JSON de configuración")
    return parser.parse_args()


# ─────────────────────────────────────────────
#  Logger
# ─────────────────────────────────────────────

def setup_logger(nombre: str, base_dir: str) -> logging.Logger:
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_dir = os.path.join(base_dir, "diagnosticos")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, f"{nombre}_{ts}.log")

    logger = logging.getLogger("extractor_01")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info(f"📝 Log: {log_path}")
    return logger


# ─────────────────────────────────────────────
#  Decodificador Price Shoes
# ─────────────────────────────────────────────

def _decodificar(txt: str, offset: int) -> str:
    resultado = []
    i = 0
    while i < len(txt):
        m = re.match(r'\(cid:(\d+)\)', txt[i:])
        if m:
            n = int(m.group(1))
            if 19 <= n <= 28:
                resultado.append(str(n - 19))
            i += len(m.group(0))
        else:
            c = txt[i]
            o = ord(c)
            if offset != 0 and 32 <= o <= 126:
                n2 = o + offset
                resultado.append(chr(n2) if 32 <= n2 <= 126 else c)
            else:
                resultado.append(c)
            i += 1
    return ''.join(resultado)


def _tiene_encoding_ps(words: list) -> bool:
    return any('(cid:' in w['text'] for w in words)


# ─────────────────────────────────────────────
#  Limpieza de precio
# ─────────────────────────────────────────────

def _limpiar_precio(val: str) -> str | None:
    if not val:
        return None
    limpio = re.sub(r'[$,\s]', '', str(val))
    limpio = re.sub(r'\.0+$', '', limpio)
    if re.match(r'^\d+(\.\d+)?$', limpio):
        return limpio.split('.')[0]
    return None


# ─────────────────────────────────────────────
#  Extractor Price Shoes
# ─────────────────────────────────────────────

class ExtractorPS:
    """
    Price Shoes: soporta AMBOS formatos del mismo proveedor:
      - PDFs con encoding propietario (cid: / offset ASCII)
      - PDFs con texto limpio y seleccionable
    La detección es automática por página.
    """

    _RE_ID = re.compile(r'^\d{6,8}$')

    def __init__(self, config: dict, logger: logging.Logger):
        self.logger   = logger
        self.offset   = config.get("encoding_offset", 29)
        self.tol_x    = config.get("tolerancia_x", 20.0)
        self.col_pag  = config.get("col_pag",    "Pag")
        self.col_id   = config.get("col_id",     "ID")
        self.col_prec = config.get("col_precio", "Sug_credito")
        self._merge_y = config.get("merge_y", 12)

    def extraer(self, pdf_path: str) -> pd.DataFrame:
        registros   = []
        col_x       = None
        ids_pdf_col = 0

        with pdfplumber.open(pdf_path) as pdf:
            self.logger.info(f"📄 Total páginas: {len(pdf.pages)}")

            for i, page in enumerate(pdf.pages, 1):
                raw = page.extract_words(x_tolerance=3, y_tolerance=5, keep_blank_chars=False)
                if not raw:
                    continue

                if _tiene_encoding_ps(raw):
                    words = [{**w, "text": _decodificar(w["text"], self.offset)} for w in raw]
                    self.logger.debug(f"  Pág {i}: encoding PS → offset={self.offset}")
                else:
                    words = raw
                    self.logger.debug(f"  Pág {i}: texto limpio")

                nuevo_cx = self._detectar_encabezado(words)
                if nuevo_cx:
                    col_x = nuevo_cx
                    self.logger.debug(
                        f"  Pág {i}: encabezado OK — "
                        f"Pag@x={col_x['pag']:.1f}  ID@x={col_x['id']:.1f}  Precio@x={col_x['precio']:.1f}"
                    )

                if not col_x:
                    self.logger.warning(f"  Pág {i}: encabezado NO encontrado — omitida")
                    continue

                x_id = col_x["id"]
                for w in words:
                    tok = w["text"].strip()
                    x0_ = w["x0"]
                    x1_ = w.get("x1", x0_)
                    near_id = (abs(x0_ - x_id) <= self.tol_x or
                               (x0_ < x_id and x_id <= x1_ + self.tol_x))
                    if not near_id:
                        continue
                    if self._RE_ID.match(tok):
                        ids_pdf_col += 1
                    else:
                        m = re.search(r'(\d{6,8})$', tok)
                        if m:
                            ids_pdf_col += 1

                filas = self._extraer_filas(words, col_x)
                antes = len(registros)
                for fila in filas:
                    registros.append({
                        "pag":         fila.get("pag", "").strip(),
                        "id":          fila.get("id",  "").strip(),
                        "precio_base": _limpiar_precio(fila.get("precio", "")),
                    })
                self.logger.debug(f"  Pág {i}: {len(registros)-antes} registros")

        df = pd.DataFrame(registros) if registros else pd.DataFrame(columns=["pag", "id", "precio_base"])
        df = self._filtrar(df)
        df.attrs["ids_esperados_pdf"] = ids_pdf_col
        return df

    def _detectar_encabezado(self, words: list) -> dict | None:
        filas_y = {}
        for w in words:
            filas_y.setdefault(round(w["top"]), []).append(w)

        ys_sorted = sorted(filas_y.keys())

        for y, tokens in sorted(filas_y.items()):
            textos = [t["text"].strip() for t in tokens]
            if self.col_pag not in textos or self.col_id not in textos:
                continue

            col_x = {}
            for t in tokens:
                txt = t["text"].strip()
                if txt == self.col_pag:
                    col_x["pag"] = t["x0"]
                elif txt == self.col_id:
                    col_x["id"] = t["x0"]
                elif txt == self.col_prec:
                    col_x["precio"] = t["x0"]

            if "precio" not in col_x:
                for y2 in ys_sorted:
                    if abs(y2 - y) == 0 or abs(y2 - y) > 5:
                        continue
                    for t in filas_y[y2]:
                        if t["text"].strip() == self.col_prec:
                            col_x["precio"] = t["x0"]
                            break

            if "pag" in col_x and "id" in col_x and "precio" in col_x:
                return col_x

        return None

    def _extraer_filas(self, words: list, col_x: dict) -> list:
        x_pag    = col_x["pag"]
        x_id     = col_x["id"]
        x_precio = col_x["precio"]
        tol      = self.tol_x

        filas_raw = {}
        for w in words:
            x     = w["x0"]
            x1_   = w.get("x1", x)
            token = w["text"].strip()
            if not token:
                continue
            if abs(x - x_pag) <= tol:
                col = "pag"
            elif (abs(x - x_id) <= tol or
                  (x < x_id and x_id <= x1_ + tol)):
                col = "id"
                if not self._RE_ID.match(token):
                    m = re.search(r'(\d{6,8})$', token)
                    if m:
                        token = m.group(1)
                    else:
                        continue
            elif abs(x - x_precio) <= tol:
                col = "precio"
            else:
                continue
            y_key = round(w["top"])
            filas_raw.setdefault(y_key, {})
            prev = filas_raw[y_key].get(col, "")
            filas_raw[y_key][col] = (prev + " " + token).strip() if prev else token

        bloques = []
        ys = sorted(filas_raw.keys())
        if not ys:
            return []

        bloque = dict(filas_raw[ys[0]])
        y_ini  = ys[0]

        for y in ys[1:]:
            if y - y_ini <= self._merge_y:
                for col, val in filas_raw[y].items():
                    if col not in bloque:
                        bloque[col] = val
                    else:
                        cur = bloque[col]
                        if col == "id" and self._RE_ID.match(val) and not self._RE_ID.match(cur):
                            bloque[col] = val
                        elif col == "precio" and re.match(r'^\$?\d', val) and not re.match(r'^\$?\d', cur):
                            bloque[col] = val
                        elif col == "pag" and re.match(r'^\d+$', val) and not re.match(r'^\d+$', cur):
                            bloque[col] = val
            else:
                bloques.append(bloque)
                bloque = dict(filas_raw[y])
                y_ini  = y

        bloques.append(bloque)
        return bloques

    def _filtrar(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        df = df[df["id"].astype(str).str.match(r"^\d{6,8}$")].copy()
        df = df[df["precio_base"].notna()].copy()
        df = df[~df["precio_base"].astype(str).str.contains(r'[a-zA-Z]', na=False)].copy()
        return df.reset_index(drop=True)


# ─────────────────────────────────────────────
#  Extractor texto limpio (Pakar / Cklass / Otro)
# ─────────────────────────────────────────────

class ExtractorTexto:

    def __init__(self, config: dict, logger: logging.Logger):
        self.logger   = logger
        self.tol_x    = config.get("tolerancia_x", 20.0)
        self.tol_col  = config.get("tol_col", 0)
        self.col_pag  = config.get("col_pag",    "PÁG.")
        self.col_id   = config.get("col_id",     "CÓDIGO")
        self.col_prec = config.get("col_precio", "2 PAGOS")
        self.offset   = config.get("encoding_offset", 0)

    def extraer(self, pdf_path: str) -> pd.DataFrame:
        registros = []
        col_x     = None

        with pdfplumber.open(pdf_path) as pdf:
            self.logger.info(f"📄 Total páginas: {len(pdf.pages)}")

            for i, page in enumerate(pdf.pages, 1):
                words = page.extract_words(x_tolerance=self.tol_x, y_tolerance=5, keep_blank_chars=False)
                if not words:
                    continue
                if self.offset:
                    words = [{**w, "text": _decodificar(w["text"], self.offset)} for w in words]

                header_y = self._detectar_encabezado(words)
                if header_y is not None:
                    col_x = self._mapear_columnas(words, header_y)
                if not col_x:
                    continue

                filas = self._agrupar_filas(words, header_y, col_x)
                for fila in filas:
                    registros.append({
                        "pag":         fila.get("pag", "").strip(),
                        "id":          fila.get("id",  "").strip(),
                        "precio_base": _limpiar_precio(fila.get("precio", "")),
                    })
                self.logger.debug(f"  Página {i}: {len(filas)} filas")

        df = pd.DataFrame(registros) if registros else pd.DataFrame(columns=["pag", "id", "precio_base"])
        return self._filtrar(df)

    def _detectar_encabezado(self, words):
        for w in words:
            txt = w["text"].strip()
            if txt == self.col_pag or self.col_pag in txt:
                return w["top"]
        return None

    def _mapear_columnas(self, words, header_y):
        col_x         = {}
        tol_y         = 15
        mapa          = {self.col_pag: "pag", self.col_id: "id", self.col_prec: "precio"}
        zona          = [w for w in words if abs(w["top"] - header_y) < tol_y]
        zona_por_fila = {}
        for w in zona:
            zona_por_fila.setdefault(round(w["top"]), []).append(w)
        for _, fila_tokens in zona_por_fila.items():
            for i, w in enumerate(fila_tokens):
                txt = w["text"].strip()
                for nombre, alias in mapa.items():
                    if alias in col_x:
                        continue
                    if txt == nombre or nombre in txt:
                        col_x[alias] = w["x0"]
                        break
                    if i + 1 < len(fila_tokens):
                        txt2 = txt + " " + fila_tokens[i + 1]["text"].strip()
                        if txt2 == nombre or nombre in txt2:
                            col_x[alias] = w["x0"]
                            break
        return col_x

    def _agrupar_filas(self, words, header_y, col_x):
        filas   = {}
        aliases = list(col_x.keys())
        xs      = [col_x[a] for a in aliases]
        for w in words:
            if w["top"] <= header_y + 5:
                continue
            y_key = round(w["top"])
            token = w["text"].strip()
            if not token:
                continue
            idx = min(range(len(xs)), key=lambda k: abs(w["x0"] - xs[k]))
            if self.tol_col > 0 and abs(w["x0"] - xs[idx]) > self.tol_col:
                continue
            col = aliases[idx]
            filas.setdefault(y_key, {})
            prev = filas[y_key].get(col, "")
            filas[y_key][col] = (prev + " " + token).strip() if prev else token

        return list(filas.values())

    def _filtrar(self, df):
        if df.empty:
            return df
        df = df[df["precio_base"].notna()].copy()
        df = df[df["id"].str.strip() != ""].copy()
        df = df[~df["precio_base"].astype(str).str.contains(r'[a-zA-Z]', na=False)].copy()
        return df.reset_index(drop=True)


# ─────────────────────────────────────────────
#  Escritura acumulativa en xlsx
# ─────────────────────────────────────────────

def _abrir_o_crear_xlsx(path: str) -> object:
    """Abre el xlsx si existe, lo crea si no."""
    if os.path.isfile(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto
    return wb


def _obtener_o_crear_pestaña(wb, nombre: str):
    """Devuelve la hoja del proveedor, la crea con encabezado si no existe."""
    if nombre in wb.sheetnames:
        return wb[nombre]
    ws = wb.create_sheet(title=nombre)
    ws.append(COLUMNAS)
    return ws


def escribir_acumulativo(df: pd.DataFrame, config: dict, tab: pd.DataFrame, logger: logging.Logger):
    """
    Abre lista_precios.xlsx, encuentra la pestaña del proveedor,
    y agrega los nuevos registros después de la última fila existente.
    """
    os.makedirs(os.path.dirname(INTEGRACION_PATH), exist_ok=True)

    proveedor  = config.get("proveedor", "Otro").strip()
    name       = config.get("name", "")
    temporada  = config.get("temporada", "")
    fecha      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = _abrir_o_crear_xlsx(INTEGRACION_PATH)
    ws = _obtener_o_crear_pestaña(wb, proveedor)

    nuevos = 0
    for record in df.itertuples(index=False):
        try:
            pb = int(float(record.precio_base))
        except (TypeError, ValueError):
            pb = record.precio_base

        pv       = _calcular_pv(pb, tab) if isinstance(pb, (int, float)) else None
        redondea = int(_round_excel(pb, -1)) if isinstance(pb, (int, float)) else None

        ws.append([
            name,        # A  name
            temporada,   # B  temporada
            record.pag,  # C  pag
            record.id,   # D  id
            pb,          # E  precio_base
            redondea,    # F  redondea
            pv,          # G  precio_venta
            fecha,       # H  fecha
        ])
        nuevos += 1

    wb.save(INTEGRACION_PATH)
    logger.info(f"✅ {nuevos} registros agregados → pestaña '{proveedor}' en {INTEGRACION_PATH}")
    return nuevos


# ─────────────────────────────────────────────
#  Clase principal
# ─────────────────────────────────────────────

class ExtractorCatalogo:

    def __init__(self, config: dict, base_dir: str):
        nombre = os.path.splitext(
            os.path.basename(config.get("_config_path", "extractor_01"))
        )[0]
        self.logger    = setup_logger(nombre, base_dir)
        self.config    = config
        self.base_dir  = base_dir
        self.proveedor = config.get("proveedor", "PS").strip()

        self.pdf_path = os.path.join(base_dir, config.get("pdf_input", ""))
        os.makedirs(os.path.dirname(INTEGRACION_PATH), exist_ok=True)

        self.logger.info(f"🏭 Proveedor:        {self.proveedor}")
        self.logger.info(f"📄 PDF de entrada:   {self.pdf_path}")
        self.logger.info(f"📊 Integración:      {INTEGRACION_PATH}  (pestaña: {self.proveedor})")
        self.logger.info(f"🏷  name:             {config.get('name', '')}")
        self.logger.info(f"📅 temporada:        {config.get('temporada', '')}")

    def _factory(self):
        p = self.proveedor.upper()
        if p == "PS":
            self.logger.info("🔧 Modo: Price Shoes (auto-detect encoding + coordenadas X)")
            return ExtractorPS(self.config, self.logger)
        elif p in ("PAKAR", "CKLASS", "OTRO"):
            self.logger.info(f"🔧 Modo: texto limpio ({self.proveedor})")
            return ExtractorTexto(self.config, self.logger)
        else:
            self.logger.warning(f"⚠️  Proveedor '{self.proveedor}' no reconocido — extractor genérico")
            return ExtractorTexto(self.config, self.logger)

    def ejecutar(self):
        if not os.path.isfile(self.pdf_path):
            self.logger.error(f"❌ PDF no encontrado: {self.pdf_path}")
            raise FileNotFoundError(self.pdf_path)

        if not os.path.isfile(TABULADOR_PATH):
            self.logger.error(f"❌ Tabulador no encontrado: {TABULADOR_PATH}")
            raise FileNotFoundError(TABULADOR_PATH)

        tab = _cargar_tabulador(TABULADOR_PATH)

        extractor = self._factory()
        self.logger.info("🚀 Iniciando extracción...")
        df = extractor.extraer(self.pdf_path)

        if df.empty:
            self.logger.warning(
                "⚠️  No se extrajeron registros.\n"
                "   Verifica col_pag y col_id en el config — deben coincidir exactamente\n"
                "   con el encabezado del PDF después de aplicar encoding_offset."
            )
            self.logger.error("🔴 EXTRACCIÓN FALLIDA")
            return

        # ── Estadísticas ──────────────────────────────────────────────────
        n         = len(df)
        u         = df["id"].nunique()
        pmin      = df["precio_base"].min()
        pmax      = df["precio_base"].max()
        esperados = df.attrs.get("ids_esperados_pdf")

        self.logger.info("─────────────────────────────────────────")
        self.logger.info(f"   Registros extraídos : {n}")
        self.logger.info(f"   IDs únicos           : {u}  {'(hay IDs repetidos)' if u < n else '(sin duplicados)'}")
        if esperados is not None:
            diff = esperados - n
            pct  = f"{n / esperados * 100:.1f}%"
            self.logger.info(f"   IDs en PDF (columna) : {esperados}")
            self.logger.info(f"   Cobertura            : {pct}  ({n}/{esperados})")
            if diff == 0:
                self.logger.info("   ✔ COMPLETO — todos los registros capturados")
            elif diff > 0:
                self.logger.warning(f"   ⚠ Faltan {diff} registro(s) — revisar manualmente")
            else:
                self.logger.warning(f"   ⚠ Extraídos {-diff} de más — posibles duplicados no filtrados")
        self.logger.info(f"   Rango precios        : ${pmin} – ${pmax}")
        self.logger.info("─────────────────────────────────────────")

        ok = esperados is None or n >= esperados * 0.98
        if n >= 100 and ok:
            self.logger.info("🟢 EXTRACCIÓN EXITOSA")
        elif n >= 20:
            self.logger.warning("🟡 EXTRACCIÓN PARCIAL — Revisar columnas en el config")
        else:
            self.logger.error("🔴 EXTRACCIÓN INSUFICIENTE — Muy pocos registros")

        # ── Acumulación en lista_precios.xlsx ────────────────────────────
        escribir_acumulativo(df, self.config, tab, self.logger)

        # ── fase_2/precios/ (sin cambios) ────────────────────────────────
        self._escribir_fase2(df, tab)

        self.logger.info(f"[STAT] proveedor={self.proveedor}")
        self.logger.info(f"[STAT] registros={n}")
        self.logger.info(f"[STAT] ids_unicos={u}")

    def _escribir_fase2(self, df: pd.DataFrame, tab: pd.DataFrame):
        """fase_2/precios/ — ID + precio_venta estáticos (sin cambios respecto a extractor.py)."""
        nombre_base = os.path.splitext(os.path.basename(self.config.get("excel_output", "salida.xlsx")))[0]
        fase2_path  = os.path.join(FASE2_PRECIOS_DIR, f"{nombre_base}.xlsx")
        os.makedirs(FASE2_PRECIOS_DIR, exist_ok=True)

        wb2  = Workbook()
        ws2  = wb2.active
        ws2.title = "Precios"
        ws2.append(["ID", "precio_venta"])

        for record in df.itertuples(index=False):
            try:
                pb = int(float(record.precio_base))
            except (TypeError, ValueError):
                pb = record.precio_base
            pv = _calcular_pv(pb, tab) if isinstance(pb, (int, float)) else None
            ws2.append([record.id, pv])

        wb2.save(fase2_path)
        self.logger.info(f"✅ Excel (fase 2) generado:  {fase2_path}")


# ─────────────────────────────────────────────
#  Punto de entrada
# ─────────────────────────────────────────────

if __name__ == "__main__":
    args        = parse_args()
    config_path = os.path.abspath(args.config)
    BASE        = os.path.dirname(os.path.abspath(__file__))

    if not os.path.isfile(config_path):
        print(f"❌ Archivo de configuración no encontrado: {config_path}")
        sys.exit(1)

    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)

    config["_config_path"] = config_path

    try:
        app = ExtractorCatalogo(config, BASE)
        app.ejecutar()
    except Exception as e:
        logging.getLogger("extractor_01").error(f"🔥 Error crítico: {e}", exc_info=True)
        sys.exit(1)
