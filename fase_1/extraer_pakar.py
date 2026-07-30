#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extraer_lista_pakar.py — Extractor de lista de precios: Pakar
Boutique Zepeda · books-label · Fase 1

Especificaciones OCR
─────────────────────────────────────────────────────────────────
Coordenadas X:
    pdfplumber extrae words con x0/x1/top.  El encabezado de columnas
    (PÁG., CÓDIGO, 2 PAGOS) calibra las posiciones X de referencia
    para toda la página.

Tolerancia horizontal (tol_x):
    Margen en px para asignar un token a su columna.
    Default: 20 px.  Configurable vía JSON.

Detección de encabezado:
    Búsqueda del token col_pag ("PÁG.") por igualdad o subcadena.
    Tolerancia Y de ±15 px para mapear las columnas.
    Soporta encabezados partidos en dos líneas adyacentes (ej. "2 PAGOS"
    como token compuesto de dos words consecutivos).

Encoding opcional:
    Si encoding_offset > 0 en el config, los tokens se decodifican
    antes de procesarse (mismo mecanismo que Price Shoes).

─────────────────────────────────────────────────────────────────
Uso:
    python3 fase_1/extraer_pakar.py --config fase_1/config/config_pakar.json

Dependencias:
    pdfplumber, pandas, openpyxl

─────────────────────────────────────────────────────────────────
Outputs
─────────────────────────────────────────────────────────────────
fase_1/salida/tabla_pakar.xlsx   (acumulativo, pestaña por config)
    catalogo | temp | pag | id | precio_base | redondea | precio_venta | fecha

fase_2/precios/<excel_output>.xlsx
    ID | precio_venta
─────────────────────────────────────────────────────────────────
"""

import math
import os
import re
import sys
import json
import logging
import argparse
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# ─────────────────────────────────────────────
#  Rutas fijas
# ─────────────────────────────────────────────

TABULADOR_PATH    = os.path.expanduser("~/books-label/fase_1/lista_cruda/tabulador.xlsx")
FASE2_PRECIOS_DIR = os.path.expanduser("~/books-label/fase_2/precios")
SALIDA_PAKAR_PATH = os.path.expanduser("~/books-label/fase_1/salida/tabla_pakar.xlsx")

COLUMNAS_STD = ["catalogo", "temp", "pag", "id", "precio_base", "redondea", "precio_venta", "fecha"]


# ─────────────────────────────────────────────
#  Helpers precio_venta
# ─────────────────────────────────────────────

def _round_excel(value: float, digits: int) -> float:
    """ROUND con redondeo 'half away from zero', igual que Excel."""
    if digits < 0:
        factor = 10 ** (-digits)
        return math.floor(float(value) / factor + 0.5) * factor
    factor = 10 ** digits
    return math.floor(float(value) * factor + 0.5) / factor


def _cargar_tabulador(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, header=0)
    df.columns = ["desde", "hasta", "sumar"]
    df = df.dropna(subset=["desde"]).copy()
    df["desde"] = pd.to_numeric(df["desde"], errors="coerce")
    df["sumar"] = pd.to_numeric(df["sumar"], errors="coerce")
    return df.sort_values("desde").reset_index(drop=True)


def _calcular_pv(precio_base, tab: pd.DataFrame):
    """Replica: =ROUND(IF(pb<200, ROUND(pb,-1)*1.5, pb+VLOOKUP(pb,tab,3,1)), -1)"""
    try:
        pb = float(precio_base)
    except (TypeError, ValueError):
        return None
    redondea = _round_excel(pb, -1)
    if pb < 200:
        return int(_round_excel(redondea * 1.5, -1))
    mask = tab["desde"] <= pb
    if not mask.any():
        return None
    sumar = float(tab.loc[mask.values.nonzero()[0][-1], "sumar"])
    return int(_round_excel(pb + sumar, -1))


# ─────────────────────────────────────────────
#  Argumentos
# ─────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Extractor de lista de precios Pakar — Boutique Zepeda"
    )
    parser.add_argument("--config", required=True, help="Ruta al archivo JSON de configuración")
    return parser.parse_args()


# ─────────────────────────────────────────────
#  Logger
# ─────────────────────────────────────────────

def setup_logger(nombre: str, base_dir: str) -> logging.Logger:
    logger = logging.getLogger("extraer_pakar")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    return logger


# ─────────────────────────────────────────────
#  Decodificador (encoding opcional)
# ─────────────────────────────────────────────

def _decodificar(txt: str, offset: int) -> str:
    resultado = []
    i = 0
    while i < len(txt):
        m = re.match(r'\(cid:(\d+)\)', txt[i:])
        if m:
            n = int(m.group(1))
            if 19 <= n <= 28:
                resultado.append(str(n - 19))
            i += len(m.group(0))
        else:
            c = txt[i]
            o = ord(c)
            if offset != 0 and 32 <= o <= 126:
                n2 = o + offset
                resultado.append(chr(n2) if 32 <= n2 <= 126 else c)
            else:
                resultado.append(c)
            i += 1
    return ''.join(resultado)


# ─────────────────────────────────────────────
#  Limpieza de precio
# ─────────────────────────────────────────────

def _limpiar_precio(val: str) -> str | None:
    """
    Normaliza cualquier formato de precio presente en los PDFs a entero:

        "$579.00"    → "579"   (decimal con punto)
        "$1,079.00"  → "1079"  (miles con coma, decimal con punto)
        "349,00"     → "349"   (decimal con coma sin miles)
        "$1.849,00"  → "1849"  (miles con punto, decimal con coma)
        "$ 69"       → "69"    (entero con espacio tras $)

    Descarta cualquier valor que contenga letras.
    """
    if not val:
        return None
    s = str(val).strip()
    if re.search(r'[a-zA-ZáéíóúÁÉÍÓÚñÑ]', s):
        return None
    s = re.sub(r'[$\s]', '', s)
    if not s:
        return None

    # Caso 0: Miles con punto Y decimal con punto: "1.159.00"
    if re.match(r'^\d{1,3}(\.\d{3})+\.\d{2}$', s):
        s = s.rsplit('.', 1)[0].replace('.', '')
    # Caso 1: Miles con punto Y decimal con coma: "1.849,00"
    elif re.match(r'^\d{1,3}(\.\d{3})+,\d{2}$', s):
        s = s.replace('.', '').split(',')[0]
    # Caso 2: Miles con coma Y decimal con punto: "1,079.00"
    elif re.match(r'^\d{1,3}(,\d{3})+\.\d+$', s):
        s = s.replace(',', '').split('.')[0]
    # Caso 3: Decimal con coma sin miles: "349,00"
    elif re.match(r'^\d+,\d{2}$', s):
        s = s.split(',')[0]
    # Caso 4: Decimal con punto: "579.00"
    elif re.match(r'^\d+\.\d+$', s):
        s = s.split('.')[0]
    # Caso 5: Entero puro: "69", "199" (ya es correcto)

    if re.match(r'^\d+$', s) and s:
        return s
    return None


# ─────────────────────────────────────────────
#  Estadísticas por página
# ─────────────────────────────────────────────

def _log_estadisticas_pagina(
    logger: logging.Logger,
    config: dict,
    registros_por_pagina: list[tuple[int, int]],
    etiqueta: str = "registros",
):
    """Log de media, desviación estándar y alertas por página."""
    if not registros_por_pagina:
        return
    conteos    = [c for _, c in registros_por_pagina]
    promedio   = sum(conteos) / len(conteos)
    desviacion = (sum((c - promedio) ** 2 for c in conteos) / len(conteos)) ** 0.5
    x          = config.get("desviacion_alerta", 2)
    alertas    = [(pag, cnt) for pag, cnt in registros_por_pagina
                  if abs(cnt - promedio) > desviacion * x]
    logger.info("═" * 45)
    logger.info(f"📊 Promedio por página: {promedio:.1f}  σ={desviacion:.1f}")
    if alertas:
        detalle = ", ".join(f"Pág {pag} ({cnt} {etiqueta})" for pag, cnt in alertas)
        logger.warning(f"  ⚠️  Páginas fuera del rango: {detalle}")
    else:
        logger.info("  ✔ Todas las páginas dentro del rango esperado")


# ─────────────────────────────────────────────
#  Extractor Pakar
# ─────────────────────────────────────────────

class ExtractorPakar:
    """Pakar: texto limpio, extracción tabular por coordenadas X."""

    def __init__(self, config: dict, logger: logging.Logger):
        self.config   = config
        self.logger   = logger
        self.tol_x    = config.get("tolerancia_x", 20.0)
        self.tol_col  = config.get("tol_col", 0)
        self.col_pag  = config.get("col_pag",    "PÁG.")
        self.col_id   = config.get("col_id",     "CÓDIGO")
        self.col_prec = config.get("col_precio", "2 PAGOS")
        self.offset   = config.get("encoding_offset", 0)

    def extraer(self, pdf_path: str) -> pd.DataFrame:
        registros = []
        col_x     = None

        registros_por_pagina = []
        with pdfplumber.open(pdf_path) as pdf:
            self.logger.info(f"📄 Total páginas: {len(pdf.pages)}")
            self.logger.info("═" * 45)

            for i, page in enumerate(pdf.pages, 1):
                words = page.extract_words(x_tolerance=self.tol_x, y_tolerance=5, keep_blank_chars=False)
                if not words:
                    continue
                if self.offset:
                    words = [{**w, "text": _decodificar(w["text"], self.offset)} for w in words]

                header_y = self._detectar_encabezado(words)
                if header_y is not None:
                    col_x = self._mapear_columnas(words, header_y)
                if not col_x:
                    continue

                filas = self._agrupar_filas(words, header_y, col_x)
                antes = len(registros)
                for fila in filas:
                    registros.append({
                        "pag":         fila.get("pag", "").strip(),
                        "id":          fila.get("id",  "").strip(),
                        "precio_base": _limpiar_precio(fila.get("precio", "")),
                    })
                df_pag = self._filtrar(pd.DataFrame(registros[antes:]))
                validos = len(df_pag)
                registros_por_pagina.append((i, validos))
                self.logger.info(f"  Página {i}: {validos} registros")

        _log_estadisticas_pagina(self.logger, self.config, registros_por_pagina, "filas")

        df = pd.DataFrame(registros) if registros else pd.DataFrame(columns=["pag", "id", "precio_base"])
        return self._filtrar(df)

    def _detectar_encabezado(self, words) -> float | None:
        for w in words:
            txt = w["text"].strip()
            if txt == self.col_pag or self.col_pag in txt:
                return w["top"]
        return None

    def _mapear_columnas(self, words, header_y) -> dict:
        col_x         = {}
        tol_y         = 15
        mapa          = {self.col_pag: "pag", self.col_id: "id", self.col_prec: "precio"}
        zona          = [w for w in words if abs(w["top"] - header_y) < tol_y]
        zona_por_fila = {}
        for w in zona:
            zona_por_fila.setdefault(round(w["top"]), []).append(w)
        for _, fila_tokens in zona_por_fila.items():
            for i, w in enumerate(fila_tokens):
                txt = w["text"].strip()
                for nombre, alias in mapa.items():
                    if alias in col_x:
                        continue
                    if txt == nombre or nombre in txt:
                        col_x[alias] = w["x0"]
                        break
                    if i + 1 < len(fila_tokens):
                        txt2 = txt + " " + fila_tokens[i + 1]["text"].strip()
                        if txt2 == nombre or nombre in txt2:
                            col_x[alias] = w["x0"]
                            break
        return col_x

    def _agrupar_filas(self, words, header_y, col_x) -> list:
        filas   = {}
        aliases = list(col_x.keys())
        xs      = [col_x[a] for a in aliases]
        for w in words:
            if w["top"] <= header_y + 5:
                continue
            y_key = round(w["top"])
            token = w["text"].strip()
            if not token:
                continue
            idx = min(range(len(xs)), key=lambda k: abs(w["x0"] - xs[k]))
            if self.tol_col > 0 and abs(w["x0"] - xs[idx]) > self.tol_col:
                continue
            col = aliases[idx]
            filas.setdefault(y_key, {})
            prev = filas[y_key].get(col, "")
            filas[y_key][col] = (prev + " " + token).strip() if prev else token
        return list(filas.values())

    def _filtrar(self, df) -> pd.DataFrame:
        if df.empty:
            return df
        df = df[df["precio_base"].notna()].copy()
        df = df[df["id"].str.strip() != ""].copy()
        df = df[~df["precio_base"].astype(str).str.contains(r'[a-zA-Z]', na=False)].copy()
        return df.reset_index(drop=True)


# ─────────────────────────────────────────────
#  Escritura acumulativa en xlsx
# ─────────────────────────────────────────────

def _abrir_o_crear_xlsx(path: str):
    if os.path.isfile(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _obtener_o_crear_pestaña(wb, nombre: str, columnas: list):
    """Devuelve la hoja; la crea con encabezado si no existe."""
    if nombre in wb.sheetnames:
        return wb[nombre]
    ws = wb.create_sheet(title=nombre)
    ws.append(columnas)
    return ws


def _estilizar_encabezado(ws):
    for cell in ws[1]:
        cell.fill      = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.font      = Font(name="Calibri", color="FFFFFF", bold=False)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")


def escribir_tabla_pakar(
    df: pd.DataFrame,
    config: dict,
    tab: pd.DataFrame,
    logger: logging.Logger,
):
    """Escritura acumulativa en tabla_pakar.xlsx (pestaña por proveedor)."""
    os.makedirs(os.path.dirname(SALIDA_PAKAR_PATH), exist_ok=True)
    proveedor = config.get("proveedor", "Pakar").strip()
    catalogo  = config.get("catalogo", "")
    temporada = config.get("temporada", "")
    fecha     = datetime.now().strftime("%Y-%m-%d")

    wb = _abrir_o_crear_xlsx(SALIDA_PAKAR_PATH)
    ws = _obtener_o_crear_pestaña(wb, proveedor, COLUMNAS_STD)

    nuevos = 0
    for record in df.itertuples(index=False):
        try:
            pb = int(float(record.precio_base))
        except (TypeError, ValueError):
            pb = record.precio_base

        pv       = _calcular_pv(pb, tab) if isinstance(pb, (int, float)) else None
        redondea = int(_round_excel(pb, -1)) if isinstance(pb, (int, float)) else None

        ws.append([catalogo, temporada, record.pag, record.id,
                   pb, redondea, pv, fecha])
        nuevos += 1

    _estilizar_encabezado(ws)
    wb.save(SALIDA_PAKAR_PATH)
    logger.info(f"✅ {nuevos} registros → '{proveedor}' en {SALIDA_PAKAR_PATH}")
    return nuevos


# ─────────────────────────────────────────────
#  Procesador principal
# ─────────────────────────────────────────────

class ProcesadorPakar:

    def __init__(self, config: dict, base_dir: str):
        nombre = os.path.splitext(
            os.path.basename(config.get("_config_path", "extractor"))
        )[0]
        self.logger    = setup_logger(nombre, base_dir)
        self.config    = config
        self.base_dir  = base_dir
        self.proveedor = config.get("proveedor", "Pakar").strip()

        self.pdf_path = os.path.join(base_dir, config.get("pdf_input", ""))
        os.makedirs(os.path.dirname(SALIDA_PAKAR_PATH), exist_ok=True)

        self.logger.info("═" * 45)
        self.logger.info(f"🏭 Proveedor:       {self.proveedor}")
        self.logger.info(f"📄 PDF de entrada:  {self.pdf_path}")
        self.logger.info(f"🛢️  Salida:          {SALIDA_PAKAR_PATH}")
        self.logger.info(f"📅 Temporada:       {config.get('temporada', '')}")

    def ejecutar(self):
        if not os.path.isfile(self.pdf_path):
            self.logger.error(f"❌ PDF no encontrado: {self.pdf_path}")
            raise FileNotFoundError(self.pdf_path)

        if not os.path.isfile(TABULADOR_PATH):
            self.logger.error(f"❌ Tabulador no encontrado: {TABULADOR_PATH}")
            raise FileNotFoundError(TABULADOR_PATH)

        tab = _cargar_tabulador(TABULADOR_PATH)
        self.logger.info("🚀 Iniciando extracción...")
        df = ExtractorPakar(self.config, self.logger).extraer(self.pdf_path)

        if df.empty:
            self.logger.warning("⚠️  No se extrajeron registros.")
            self.logger.error("🔴 EXTRACCIÓN FALLIDA")
            return

        # ── Estadísticas generales ────────────────────────────────────
        n    = len(df)
        pmin = df["precio_base"].min()
        pmax = df["precio_base"].max()
        u    = df["id"].nunique()

        self.logger.info("═" * 45)
        self.logger.info(f"   Registros extraídos  : {n}")
        self.logger.info(f"   IDs únicos           : {u}  {'(hay IDs repetidos)' if u < n else '(sin duplicados)'}")
        self.logger.info(f"   Rango precios        : ${pmin} – ${pmax}")
        self.logger.info("═" * 45)

        # ── fase_1/salida/tabla_pakar.xlsx ────────────────────────────
        escribir_tabla_pakar(df, self.config, tab, self.logger)

        # ── fase_2/precios/ ───────────────────────────────────────────
        self._escribir_fase2(df, tab)

        self.logger.info(f"[STAT] proveedor={self.proveedor}")
        self.logger.info(f"[STAT] registros={n}")

    def _escribir_fase2(self, df: pd.DataFrame, tab: pd.DataFrame):
        nombre_base = os.path.splitext(
            os.path.basename(self.config.get("excel_output", "salida.xlsx"))
        )[0]
        fase2_path = os.path.join(FASE2_PRECIOS_DIR, f"{nombre_base}.xlsx")
        os.makedirs(FASE2_PRECIOS_DIR, exist_ok=True)

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Precios"

        ws2.append(["ID", "precio_venta"])
        for record in df.itertuples(index=False):
            try:
                pb = int(float(record.precio_base))
            except (TypeError, ValueError):
                pb = record.precio_base
            pv = _calcular_pv(pb, tab) if isinstance(pb, (int, float)) else None
            ws2.append([record.id, pv])

        wb2.save(fase2_path)
        self.logger.info(f"✅ Excel (fase 2) generado: {fase2_path}")


# ─────────────────────────────────────────────
#  Punto de entrada
# ─────────────────────────────────────────────

if __name__ == "__main__":
    args        = parse_args()
    config_path = os.path.abspath(args.config)
    BASE        = os.path.dirname(os.path.abspath(__file__))

    if not os.path.isfile(config_path):
        print(f"❌ Archivo de configuración no encontrado: {config_path}")
        sys.exit(1)

    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)

    config["_config_path"] = config_path

    try:
        ProcesadorPakar(config, BASE).ejecutar()
    except Exception as e:
        logging.getLogger("extraer_pakar").error(f"🔥 Error crítico: {e}", exc_info=True)
        sys.exit(1)
