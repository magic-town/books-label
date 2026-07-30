#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extraer_lista_price.py — Extractor de lista de precios: Price Shoes
Boutique Zepeda · books-label · Fase 1

Especificaciones OCR
─────────────────────────────────────────────────────────────────
Coordenadas X:
    pdfplumber extrae words con x0/x1/top.  El encabezado de columnas
    (Pag, ID, Sug_credito) calibra las posiciones X de referencia
    para toda la página.

Tolerancia horizontal (tol_x):
    Margen en px para asignar un token a su columna.
    Default: 20 px.  Configurable vía JSON.

Merge vertical (merge_y):
    Tokens de la misma fila lógica pueden tener tops separados hasta
    16 px por fragmentación del glifo.  Cuando dos sub-filas caen
    dentro de merge_y px, se fusionan priorizando:
      - ID: token que matchea ^\d{5,8}$
      - Precio: token que empieza con $\d
      - Pag: token numérico puro

Encoding propietario:
    Algunos PDFs codifican dígitos como (cid:19)–(cid:28) → 0–9 y
    desplazan ASCII +offset (default 29).  La detección es automática
    por página.

Regex ID:  ^\d{5,8}$
    Mínimo 5 dígitos para capturar IDs de marcas importadas y
    accesorios con IDs cortos (Converse, Vans, K-Swiss…).

Filtro pag numérico:
    Descarta filas donde `pag` contiene texto (encabezados / pies de
    página que se cuelan como registros).

Detección de encabezado:
    Tolerancia Y de ±15 px al buscar Sug_credito.
    Fallback: cualquier token que contenga la subcadena de col_precio.

Formato precio PS México:
    Miles.decimal.centavos ("$1.159.00" → 1159).

─────────────────────────────────────────────────────────────────
Uso:
    python3 fase_1/extraer_price.py --config fase_1/config/config_price.json

Dependencias:
    pdfplumber, pandas, openpyxl

─────────────────────────────────────────────────────────────────
Outputs
─────────────────────────────────────────────────────────────────
fase_1/salida/tabla_price.xlsx   (acumulativo, pestaña por proveedor)
    id | catalogo | temp | pag | marca | corrida | claves | precio_base |
    contado_antes | contado_despues | precio_venta | fecha

    `claves` se extrae del PDF, se normaliza a numérico y se oculta en
    el spreadsheet (se usa solo como base de cálculo).

    `precio_base`, `contado_antes`, `contado_despues` y `precio_venta`
    ya NO se extraen del PDF: se calculan a partir de `claves` mediante
    tablas de incrementos fijos (ver TABLA_* más abajo).

fase_2/precios/<excel_output>.xlsx
    ID | precio_venta
─────────────────────────────────────────────────────────────────
"""

import math
import os
import re
import sys
import json
import logging
import argparse
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  Rutas fijas
# ─────────────────────────────────────────────

FASE2_PRECIOS_DIR = os.path.expanduser("~/books-label/fase_2/precios")
SALIDA_PRICE_PATH = os.path.expanduser("~/books-label/fase_1/salida/tabla_price.xlsx")

COLUMNAS_STD = [
    "id", "catalogo", "temp", "pag", "marca", "corrida", "claves",
    "precio_base", "contado_antes", "contado_despues", "precio_venta",
    "fecha",
]

# Columna que se oculta en el spreadsheet de salida (se usa solo como
# base de cálculo, no necesita mostrarse).
COLUMNAS_OCULTAS = {"claves"}

# Columnas cuyo valor puede venir partido en 2+ renglones dentro del
# mismo bloque de fila (p.ej. "Marca" ocupa 2 líneas) y por lo tanto
# deben concatenarse en vez de conservar solo el primer valor visto.
COLUMNAS_MULTILINEA = {"marca"}


# ─────────────────────────────────────────────
#  Helpers de cálculo (precio_base, contado_antes/despues, precio_venta)
# ─────────────────────────────────────────────
#
# Todas las tablas usan rangos [desde, hasta] inclusivos; el último
# rango (desde=2001, hasta=None) cubre "2001 en adelante".

# precio_base = claves + % según rango de `claves`
TABLA_PRECIO_BASE_PCT = [
    (1,    170,  0.50),
    (171,  600,  0.55),
    (601,  1000, 0.58),
    (1001, 1500, 0.58),
    (1501, 2000, 0.60),
    (2001, None, 0.60),
]

# contado_antes = precio_base + monto fijo según rango de `precio_base`
TABLA_CONTADO_ANTES_ADD = [
    (1,    170,  10),
    (171,  600,  20),
    (601,  1000, 40),
    (1001, 1500, 50),
    (1501, 2000, 70),
    (2001, None, 80),
]

# contado_despues = precio_base + monto fijo según rango de `precio_base`
TABLA_CONTADO_DESPUES_ADD = [
    (1,    170,  30),
    (171,  600,  40),
    (601,  1000, 80),
    (1001, 1500, 100),
    (1501, 2000, 120),
    (2001, None, 120),
]

# precio_venta = claves + % según rango de `claves`
TABLA_PRECIO_VENTA_PCT = [
    (1,    170,  0.75),
    (171,  600,  0.78),
    (601,  1000, 0.80),
    (1001, 1500, 0.83),
    (1501, 2000, 0.86),
    (2001, None, 0.90),
]


def _round_excel(value: float, digits: int) -> float:
    """ROUND con redondeo 'half away from zero', igual que Excel.

    Con digits=-1 esto implementa exactamente la regla del negocio:
    unidad < 5 → redondea al decimal inferior; unidad >= 5 → redondea
    al decimal superior.
    """
    if digits < 0:
        factor = 10 ** (-digits)
        return math.floor(float(value) / factor + 0.5) * factor
    factor = 10 ** digits
    return math.floor(float(value) * factor + 0.5) / factor


def _buscar_en_tabla(valor: float, tabla: list):
    """Devuelve el dato (% o monto) cuyo rango [desde, hasta] contiene
    `valor`. `hasta=None` significa "en adelante".

    Selecciona el tramo por el mayor `desde` que sea <= valor (equivalente
    a un VLOOKUP aproximado), en vez de depender del orden de iteración
    de la lista, para evitar cualquier ambigüedad en los límites
    compartidos entre tramos (p.ej. 600/601).
    """
    mejor = None
    for desde, hasta, dato in tabla:
        if valor < desde:
            continue
        if hasta is not None and valor > hasta:
            continue
        if mejor is None or desde > mejor[0]:
            mejor = (desde, dato)
    return mejor[1] if mejor is not None else None


# ─────────────────────────────────────────────
#  Argumentos
# ─────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Extractor de lista de precios Price Shoes — Boutique Zepeda"
    )
    parser.add_argument("--config", required=True, help="Ruta al archivo JSON de configuración")
    return parser.parse_args()


# ─────────────────────────────────────────────
#  Logger
# ─────────────────────────────────────────────

def setup_logger(nombre: str, base_dir: str) -> logging.Logger:
    logger = logging.getLogger("extraer_price")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    return logger


# ─────────────────────────────────────────────
#  Decodificador Price Shoes
# ─────────────────────────────────────────────

def _decodificar(txt: str, offset: int) -> str:
    resultado = []
    i = 0
    while i < len(txt):
        m = re.match(r'\(cid:(\d+)\)', txt[i:])
        if m:
            n = int(m.group(1))
            if 19 <= n <= 28:
                resultado.append(str(n - 19))
            i += len(m.group(0))
        else:
            c = txt[i]
            o = ord(c)
            if offset != 0 and 32 <= o <= 126:
                n2 = o + offset
                resultado.append(chr(n2) if 32 <= n2 <= 126 else c)
            else:
                resultado.append(c)
            i += 1
    return ''.join(resultado)


def _tiene_encoding_ps(words: list) -> bool:
    return any('(cid:' in w['text'] for w in words)


# ─────────────────────────────────────────────
#  Extracción y normalización de `claves`
# ─────────────────────────────────────────────

def _extraer_claves(val: str) -> int | None:
    """
    Formato en PDF: dígito(s) + separador no numérico + dígito(s)
    (p.ej. "1PS23", "'123", "'1234"), de 3 o 4 dígitos en total.
    Regla: eliminar todos los caracteres no numéricos y concatenar los
    dígitos restantes en su orden original, devolviendo un entero.

        "1PS23"  → 123
        "1PS234" → 1234
        "'123"   → 123

    `claves` debe ser numérico (no texto) porque se usa para calcular
    precio_base, contado_antes, contado_despues y precio_venta.
    """
    if not val:
        return None
    digitos = re.sub(r'[^0-9]', '', str(val))
    if not digitos:
        return None
    return int(digitos)


# ─────────────────────────────────────────────
#  Estadísticas por página
# ─────────────────────────────────────────────

def _log_estadisticas_pagina(
    logger: logging.Logger,
    config: dict,
    registros_por_pagina: list[tuple[int, int]],
    etiqueta: str = "registros",
):
    """Log de media, desviación estándar y alertas por página."""
    if not registros_por_pagina:
        return
    conteos    = [c for _, c in registros_por_pagina]
    promedio   = sum(conteos) / len(conteos)
    desviacion = (sum((c - promedio) ** 2 for c in conteos) / len(conteos)) ** 0.5
    x          = config.get("desviacion_alerta", 2)
    alertas    = [(pag, cnt) for pag, cnt in registros_por_pagina
                  if abs(cnt - promedio) > desviacion * x]
    logger.info("═" * 45)
    logger.info(f"📊 Promedio por página: {promedio:.1f}  σ={desviacion:.1f}")
    if alertas:
        detalle = ", ".join(f"Pág {pag} ({cnt} {etiqueta})" for pag, cnt in alertas)
        logger.warning(f"  ⚠️  Páginas fuera del rango: {detalle}")
    else:
        logger.info("  ✔ Todas las páginas dentro del rango esperado")


# ─────────────────────────────────────────────
#  Extractor Price Shoes
# ─────────────────────────────────────────────

class ExtractorPS:
    """
    Price Shoes: soporta AMBOS formatos del mismo proveedor:
      - PDFs con encoding propietario (cid: / offset ASCII)
      - PDFs con texto limpio y seleccionable
    La detección es automática por página.
    """

    # Mínimo 5 dígitos para capturar IDs de marcas importadas
    # (Converse: 10102, 72343; Vans: 89807…) y accesorios PS (96395, 63526…)
    _RE_ID = re.compile(r'^\d{5,8}$')

    def __init__(self, config: dict, logger: logging.Logger):
        self.config   = config
        self.logger   = logger
        self.offset   = config.get("encoding_offset", 29)
        self.tol_x    = config.get("tolerancia_x", 20.0)

        # Etiquetas de encabezado tal como aparecen en el PDF, mapeadas
        # al nombre interno de cada columna. "pag" e "id" son las
        # columnas ancla usadas para localizar la fila de encabezado.
        self.col_labels = {
            "pag":     config.get("col_pag",     "Pag"),
            "id":      config.get("col_id",      "ID"),
            "marca":   config.get("col_marca",   "Marca"),
            "corrida": config.get("col_corrida", "Corrida"),
            "claves":  config.get("col_claves",  "Claves"),
        }
        # Tolerancia vertical para fusionar sub-filas fragmentadas
        # (marcas partidas en múltiples líneas).
        self._merge_y = config.get("merge_y", 16)

    def extraer(self, pdf_path: str) -> pd.DataFrame:
        registros   = []
        col_x       = None
        ids_pdf_col = 0

        registros_por_pagina = []
        with pdfplumber.open(pdf_path) as pdf:
            self.logger.info(f"📄 Total páginas: {len(pdf.pages)}")
            self.logger.info("═" * 45)

            for i, page in enumerate(pdf.pages, 1):
                raw = page.extract_words(x_tolerance=3, y_tolerance=5, keep_blank_chars=False)
                if not raw:
                    self.logger.info(f"  Pág {i}: sin palabras — omitida")
                    continue

                if _tiene_encoding_ps(raw):
                    words = [{**w, "text": _decodificar(w["text"], self.offset)} for w in raw]
                    self.logger.debug(f"  Pág {i}: encoding PS → offset={self.offset}")
                else:
                    words = raw
                    self.logger.debug(f"  Pág {i}: texto limpio")

                nuevo_cx = self._detectar_encabezado(words)
                if nuevo_cx:
                    col_x = nuevo_cx
                    resumen = "  ".join(f"{k}@x={v:.1f}" for k, v in col_x.items())
                    self.logger.debug(f"  Pág {i}: encabezado OK — {resumen}")
                else:
                    self.logger.debug(f"  Pág {i}: encabezado no encontrado en esta página")

                if not col_x:
                    self.logger.warning(f"  Pág {i}: encabezado NO encontrado — omitida")
                    continue

                x_id = col_x["id"]
                for w in words:
                    tok = w["text"].strip()
                    x0_ = w["x0"]
                    x1_ = w.get("x1", x0_)
                    near_id = (abs(x0_ - x_id) <= self.tol_x or
                               (x0_ < x_id and x_id <= x1_ + self.tol_x))
                    if not near_id:
                        continue
                    if self._RE_ID.match(tok):
                        ids_pdf_col += 1
                    else:
                        m = re.search(r'(\d{5,8})$', tok)
                        if m:
                            ids_pdf_col += 1

                filas = self._extraer_filas(words, col_x)
                antes = len(registros)
                for fila in filas:
                    registros.append({
                        "pag":     fila.get("pag", "").strip(),
                        "id":      fila.get("id",  "").strip(),
                        "marca":   fila.get("marca", "").strip(),
                        "corrida": fila.get("corrida", "").strip(),
                        # claves: numérico, base de cálculo de precio_base,
                        # contado_antes, contado_despues y precio_venta.
                        "claves":  _extraer_claves(fila.get("claves", "")),
                    })
                df_pag = self._filtrar(pd.DataFrame(registros[antes:]))
                validos = len(df_pag)
                registros_por_pagina.append((i, validos))
                self.logger.info(f"  Pág {i}: {validos} registros")

        _log_estadisticas_pagina(self.logger, self.config, registros_por_pagina)

        df = pd.DataFrame(registros) if registros else pd.DataFrame(columns=[
            "pag", "id", "marca", "corrida", "claves"
        ])
        df = self._filtrar(df)
        df.attrs["ids_esperados_pdf"] = ids_pdf_col
        return df

    # Columnas obligatorias: sin ellas no se acepta la fila como encabezado.
    _COLS_OBLIGATORIAS = ("pag", "id", "claves")

    def _buscar_columna(self, label: str, y: int, filas_y: dict, ys_sorted: list) -> float | None:
        """Busca `label` en la fila y, con tolerancia ±15px en Y, y como
        último recurso por subcadena (variaciones tipográficas)."""
        for t in filas_y.get(y, []):
            if t["text"].strip() == label:
                return t["x0"]
        for y2 in ys_sorted:
            if y2 == y or abs(y2 - y) > 15:
                continue
            for t in filas_y[y2]:
                if t["text"].strip() == label:
                    return t["x0"]
        label_lower = label.lower()
        for y2 in ys_sorted:
            if abs(y2 - y) > 15:
                continue
            for t in filas_y[y2]:
                if label_lower in t["text"].lower():
                    return t["x0"]
        return None

    def _detectar_encabezado(self, words: list) -> dict | None:
        """
        Detecta la fila de encabezado y devuelve, para cada columna
        configurada (pag, id, marca, corrida, claves), la coordenada X
        de referencia.

        "pag" e "id" en la misma fila anclan la búsqueda del encabezado.
        pag/id/claves son obligatorias; el resto se agrega si se
        encuentra (tolerancia Y ±15px, con fallback por subcadena).
        """
        filas_y = {}
        for w in words:
            filas_y.setdefault(round(w["top"]), []).append(w)

        ys_sorted = sorted(filas_y.keys())

        for y, tokens in sorted(filas_y.items()):
            textos = [t["text"].strip() for t in tokens]
            if self.col_labels["pag"] not in textos or self.col_labels["id"] not in textos:
                continue

            col_x = {}
            for nombre, label in self.col_labels.items():
                x = self._buscar_columna(label, y, filas_y, ys_sorted)
                if x is not None:
                    col_x[nombre] = x

            if all(c in col_x for c in self._COLS_OBLIGATORIAS):
                faltantes = [c for c in self.col_labels if c not in col_x]
                if faltantes:
                    self.logger.warning(f"  ⚠️  Encabezado sin columnas: {', '.join(faltantes)}")
                return col_x

        return None

    def _extraer_filas(self, words: list, col_x: dict) -> list:
        tol = self.tol_x
        x_id = col_x["id"]

        filas_raw = {}
        for w in words:
            x     = w["x0"]
            x1_   = w.get("x1", x)
            token = w["text"].strip()
            if not token:
                continue

            # Columna más cercana entre las configuradas para esta página.
            col, mejor_dist = None, None
            for nombre, x_col in col_x.items():
                dist = abs(x - x_col)
                if dist <= tol and (mejor_dist is None or dist < mejor_dist):
                    col, mejor_dist = nombre, dist
            if col is None and x < x_id and x_id <= x1_ + tol:
                col = "id"
            if col is None:
                continue

            if col == "id":
                if not self._RE_ID.match(token):
                    m = re.search(r'(\d{5,8})$', token)
                    if m:
                        token = m.group(1)
                    else:
                        continue

            y_key = round(w["top"])
            filas_raw.setdefault(y_key, {})
            prev = filas_raw[y_key].get(col, "")
            filas_raw[y_key][col] = (prev + " " + token).strip() if prev else token

        bloques = []
        ys = sorted(filas_raw.keys())
        if not ys:
            return []

        bloque = dict(filas_raw[ys[0]])
        y_ini  = ys[0]

        for y in ys[1:]:
            if y - y_ini <= self._merge_y:
                for col, val in filas_raw[y].items():
                    if col not in bloque:
                        bloque[col] = val
                    else:
                        cur = bloque[col]
                        if col in COLUMNAS_MULTILINEA:
                            bloque[col] = (cur + " " + val).strip()
                        elif col == "id" and self._RE_ID.match(val) and not self._RE_ID.match(cur):
                            bloque[col] = val
                        elif col == "claves" and re.match(r'^\W*\d', val) and not re.match(r'^\W*\d', cur):
                            bloque[col] = val
                        elif col == "pag" and re.match(r'^\d+$', val) and not re.match(r'^\d+$', cur):
                            bloque[col] = val
            else:
                bloques.append(bloque)
                bloque = dict(filas_raw[y])
                y_ini  = y

        bloques.append(bloque)
        return bloques

    def _filtrar(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        # Solo IDs con formato numérico 5-8 dígitos
        df = df[df["id"].astype(str).str.match(r"^\d{5,8}$")].copy()
        # Solo filas con `claves` numérico válido (base de todos los cálculos)
        df = df[df["claves"].notna()].copy()
        # Descartar filas donde pag contiene texto (encabezado/pie de página)
        df = df[df["pag"].astype(str).str.match(r"^\d+$")].copy()
        return df.reset_index(drop=True)


# ─────────────────────────────────────────────
#  Escritura acumulativa en xlsx
# ─────────────────────────────────────────────

def _abrir_o_crear_xlsx(path: str):
    if os.path.isfile(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _obtener_o_crear_pestaña(wb, nombre: str, columnas: list):
    """Devuelve la hoja; la crea con encabezado si no existe."""
    if nombre in wb.sheetnames:
        return wb[nombre]
    ws = wb.create_sheet(title=nombre)
    ws.append(columnas)
    return ws


def _estilizar_encabezado(ws):
    for cell in ws[1]:
        cell.fill      = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.font      = Font(name="Calibri", color="FFFFFF", bold=False)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")


def _calcular_derivados(record) -> dict:
    """Calcula precio_base, contado_antes, contado_despues y precio_venta
    a partir de `claves`, usando las tablas de incrementos fijos.

    - precio_base    = round10(claves * (1 + %precio_base(claves)))
    - contado_antes  = precio_base + monto(precio_base)
    - contado_despues= precio_base + monto(precio_base)
    - precio_venta   = round10(claves * (1 + %precio_venta(claves)))
    """
    claves = getattr(record, "claves", None)
    try:
        claves_num = int(float(claves))
    except (TypeError, ValueError):
        return {
            "precio_base": None, "contado_antes": None,
            "contado_despues": None, "precio_venta": None,
        }

    pct_pb = _buscar_en_tabla(claves_num, TABLA_PRECIO_BASE_PCT)
    precio_base = (
        int(_round_excel(claves_num * (1 + pct_pb), -1)) if pct_pb is not None else None
    )

    contado_antes = contado_despues = None
    if precio_base is not None:
        add_antes = _buscar_en_tabla(precio_base, TABLA_CONTADO_ANTES_ADD)
        add_despues = _buscar_en_tabla(precio_base, TABLA_CONTADO_DESPUES_ADD)
        if add_antes is not None:
            contado_antes = precio_base + add_antes
        if add_despues is not None:
            contado_despues = precio_base + add_despues

    pct_pv = _buscar_en_tabla(claves_num, TABLA_PRECIO_VENTA_PCT)
    precio_venta = (
        int(_round_excel(claves_num * (1 + pct_pv), -1)) if pct_pv is not None else None
    )

    return {
        "precio_base": precio_base,
        "contado_antes": contado_antes,
        "contado_despues": contado_despues,
        "precio_venta": precio_venta,
    }


def escribir_tabla_price(
    df: pd.DataFrame,
    config: dict,
    logger: logging.Logger,
):
    """Escritura acumulativa en tabla_price.xlsx (pestaña por proveedor)."""
    os.makedirs(os.path.dirname(SALIDA_PRICE_PATH), exist_ok=True)
    proveedor = config.get("proveedor", "PS").strip()
    catalogo  = config.get("catalogo", "")
    temporada = config.get("temporada", "")
    fecha     = datetime.now().strftime("%Y-%m-%d")

    wb = _abrir_o_crear_xlsx(SALIDA_PRICE_PATH)
    ws = _obtener_o_crear_pestaña(wb, proveedor, COLUMNAS_STD)

    nuevos = 0
    for record in df.itertuples(index=False):
        d = _calcular_derivados(record)

        claves_val = getattr(record, "claves", None)
        try:
            claves_val = int(float(claves_val)) if claves_val not in (None, "") else None
        except (TypeError, ValueError):
            claves_val = None

        ws.append([
            record.id, catalogo, temporada, record.pag,
            getattr(record, "marca", None), getattr(record, "corrida", None),
            claves_val,
            d["precio_base"], d["contado_antes"], d["contado_despues"], d["precio_venta"],
            fecha,
        ])
        nuevos += 1

    _ocultar_columnas(ws)
    _estilizar_encabezado(ws)
    wb.save(SALIDA_PRICE_PATH)
    logger.info(f"✅ {nuevos} registros → '{proveedor}' en {SALIDA_PRICE_PATH}")
    return nuevos


def _ocultar_columnas(ws):
    """Oculta en el spreadsheet las columnas listadas en COLUMNAS_OCULTAS
    (p.ej. `claves`, que se conserva solo como base de cálculo)."""
    for nombre in COLUMNAS_OCULTAS:
        if nombre not in COLUMNAS_STD:
            continue
        idx = COLUMNAS_STD.index(nombre) + 1  # 1-based
        letra = get_column_letter(idx)
        ws.column_dimensions[letra].hidden = True


# ─────────────────────────────────────────────
#  Procesador principal
# ─────────────────────────────────────────────

class ProcesadorPrice:

    def __init__(self, config: dict, base_dir: str):
        nombre = os.path.splitext(
            os.path.basename(config.get("_config_path", "extractor"))
        )[0]
        self.logger    = setup_logger(nombre, base_dir)
        self.config    = config
        self.base_dir  = base_dir
        self.proveedor = config.get("proveedor", "PS").strip()

        self.pdf_path = os.path.join(base_dir, config.get("pdf_input", ""))
        os.makedirs(os.path.dirname(SALIDA_PRICE_PATH), exist_ok=True)

        self.logger.info("═" * 45)
        self.logger.info(f"🏭 Proveedor:       {self.proveedor}")
        self.logger.info(f"📄 PDF de entrada:  {self.pdf_path}")
        self.logger.info(f"🛢️  Salida:          {SALIDA_PRICE_PATH}")
        self.logger.info(f"📅 Temporada:       {config.get('temporada', '')}")

    def ejecutar(self):
        if not os.path.isfile(self.pdf_path):
            self.logger.error(f"❌ PDF no encontrado: {self.pdf_path}")
            raise FileNotFoundError(self.pdf_path)

        self.logger.info("🚀 Iniciando extracción...")
        df = ExtractorPS(self.config, self.logger).extraer(self.pdf_path)

        if df.empty:
            self.logger.warning("⚠️  No se extrajeron registros.")
            self.logger.error("🔴 EXTRACCIÓN FALLIDA")
            return

        # ── Estadísticas generales ────────────────────────────────────
        n         = len(df)
        u         = df["id"].nunique()
        esperados = df.attrs.get("ids_esperados_pdf")

        precios_base = [
            d["precio_base"] for d in (
                _calcular_derivados(r) for r in df.itertuples(index=False)
            ) if d["precio_base"] is not None
        ]
        pmin = min(precios_base) if precios_base else None
        pmax = max(precios_base) if precios_base else None

        self.logger.info("═" * 45)
        self.logger.info(f"   Registros extraídos  : {n}")
        self.logger.info(f"   IDs únicos           : {u}  {'(hay IDs repetidos)' if u < n else '(sin duplicados)'}")
        if esperados is not None:
            diff = esperados - n
            pct  = f"{n / esperados * 100:.1f}%"
            self.logger.info(f"   IDs en PDF (columna) : {esperados}")
            self.logger.info(f"   Cobertura            : {pct}  ({n}/{esperados})")
            if diff == 0:
                self.logger.info("   ✔ COMPLETO — todos los registros capturados")
            elif diff > 0:
                self.logger.warning(f"   ⚠ Faltan {diff} registro(s) — revisar manualmente")
            else:
                self.logger.warning(f"   ⚠ Extraídos {-diff} de más — posibles duplicados")
        self.logger.info(f"   Rango precio_base    : ${pmin} – ${pmax}")
        self.logger.info("═" * 45)

        # ── fase_1/salida/tabla_price.xlsx ────────────────────────────
        escribir_tabla_price(df, self.config, self.logger)

        # ── fase_2/precios/ ───────────────────────────────────────────
        self._escribir_fase2(df)

        self.logger.info(f"[STAT] proveedor={self.proveedor}")
        self.logger.info(f"[STAT] registros={n}")

    def _escribir_fase2(self, df: pd.DataFrame):
        nombre_base = os.path.splitext(
            os.path.basename(self.config.get("excel_output", "salida.xlsx"))
        )[0]
        fase2_path = os.path.join(FASE2_PRECIOS_DIR, f"{nombre_base}.xlsx")
        os.makedirs(FASE2_PRECIOS_DIR, exist_ok=True)

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Precios"

        ws2.append(["ID", "precio_venta"])
        for record in df.itertuples(index=False):
            d = _calcular_derivados(record)
            ws2.append([record.id, d["precio_venta"]])

        wb2.save(fase2_path)
        self.logger.info(f"✅ Excel (fase 2) generado: {fase2_path}")


# ─────────────────────────────────────────────
#  Punto de entrada
# ─────────────────────────────────────────────

if __name__ == "__main__":
    args        = parse_args()
    config_path = os.path.abspath(args.config)
    BASE        = os.path.dirname(os.path.abspath(__file__))

    if not os.path.isfile(config_path):
        print(f"❌ Archivo de configuración no encontrado: {config_path}")
        sys.exit(1)

    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)

    config["_config_path"] = config_path

    try:
        ProcesadorPrice(config, BASE).ejecutar()
    except Exception as e:
        logging.getLogger("extraer_price").error(f"🔥 Error crítico: {e}", exc_info=True)
        sys.exit(1)